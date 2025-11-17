# -*- coding: utf-8 -*-
"""
委託販売精算書を生成するためのモジュール
Excelファイルを操作して、顧客ごとの精算書シートを作成します。
"""

# --- 必要なライブラリのインポート ---
# openpyxl: Excelファイルを読み書きするためのライブラリ
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# datetime: 日付を扱うためのモジュール
from datetime import date

# calendar: カレンダー関連の機能（例：月の最終日を取得）を提供
import calendar

# os: ファイルパスの操作など、OS関連の機能を提供
import os

# typing: 型ヒント（変数の型を明示する）に使用
from typing import List, Dict, Any

# pandas: データ分析ライブラリ。データの集計や変換に使用
import pandas as pd


class SettlementGenerator:
    """
    委託販売精算書を生成するクラス

    このクラスは、テンプレートシートをコピーして顧客ごとの精算書シートを作成します。
    売上データを集計し、手数料や消費税を計算して、Excelファイルに書き込みます。
    """

    def __init__(self):
        """
        クラスの初期化（コンストラクタ）

        現在は特に初期化処理は行いませんが、将来的に設定を保存する場合に使用します。
        """
        pass

    def create_settlement_sheet(
        self,
        wb: Workbook,
        template_sheet: Worksheet,
        customer_data: Dict[str, Any],
        sales_data: List[Dict[str, Any]],
        year: int,
        month: int,
        bank_transfer_fee: int = 440,
    ):
        """
        既存のExcelブックに、顧客一社分の精算書シートを追加するメイン関数

        Args:
            wb: 精算書を追加するExcelブック（Workbookオブジェクト）
            template_sheet: コピー元となるテンプレートシート
            customer_data: 顧客情報が入った辞書（会社名、住所、銀行口座情報など）
            sales_data: 売上明細のリスト（各要素は商品コード、商品名、単価、販売数、売上金額など）
            year: 精算対象の年
            month: 精算対象の月
            bank_transfer_fee: 振込手数料（デフォルトは440円）
        """
        # --- シートの作成とタイトル設定 ---
        # 顧客データから会社名を取得（なければ「不明な顧客」）
        client_name = customer_data.get("会社名", "不明な顧客")
        # テンプレートシートをコピーして新しいシートを作成
        ws = wb.copy_worksheet(template_sheet)
        # Excelのシート名は最大31文字までなので、それ以上は切り詰める
        ws.title = client_name[:31]

        # --- 日付の計算 ---
        # calendar.monthrange(): 指定した年月の最初の曜日と最終日を取得
        # 例: (0, 31) → 0は月曜日、31はその月の最終日
        _, last_day = calendar.monthrange(year, month)
        # 精算期間の開始日: その月の1日
        period_start = date(year, month, 1)
        # 精算期間の終了日: その月の最終日
        period_end = date(year, month, last_day)
        # 発行日: 精算対象月の翌月25日（12月の場合は翌年1月25日）
        issue_date = date(year, month + 1, 25) if month < 12 else date(year + 1, 1, 25)
        # 振込予定日: 精算対象月の翌々月10日（11月・12月の場合は適切に調整）
        due_date = date(year, month + 2, 10) if month < 11 else date(year + 1, (month + 2) % 12, 10)
        # 顧客IDを取得（なければ「N/A」）
        client_id = customer_data.get("クライアントID", "N/A")
        # 精算番号を生成（例: "2025-10-C001"）
        settlement_number = f"{year}-{month:02d}-{client_id}"

        # --- 売上データの集計処理 ---
        # 同じ商品コードの売上をまとめて集計します（例：商品Aが複数回売れた場合、販売数を合計）
        if not sales_data:
            # 売上データが空の場合は、空のリストと0を設定
            aggregated_sales = []
            total_sales = 0
        else:
            # pandasのDataFrame（表形式のデータ）に変換
            # これにより、データの集計や変換が簡単になります
            sales_df = pd.DataFrame(sales_data)

            # データ型を数値に変換（文字列で入っている可能性があるため）
            # pd.to_numeric(): 文字列を数値に変換
            # errors="coerce": 変換できない値はNaN（欠損値）にする
            # fillna(0): NaNを0に置き換える
            sales_df["販売数"] = pd.to_numeric(sales_df["販売数"], errors="coerce").fillna(0)
            sales_df["売上金額"] = pd.to_numeric(sales_df["売上金額"], errors="coerce").fillna(0)
            sales_df["単価"] = pd.to_numeric(sales_df["単価"], errors="coerce").fillna(0)

            # 商品コードでグループ化して集計
            # groupby("商品コード"): 同じ商品コードの行をグループ化
            # .agg(): 各グループに対して集計処理を実行
            #   - 商品名: 最初の値を取得（同じ商品コードなら商品名も同じはず）
            #   - 単価: 最初の値を取得
            #   - 販売数: 合計を計算
            #   - 売上金額: 合計を計算
            # .reset_index(): インデックスをリセットして、通常の列に戻す
            aggregated_df = (
                sales_df.groupby("商品コード")
                .agg(
                    商品名=("商品名", "first"),
                    単価=("単価", "first"),
                    販売数=("販売数", "sum"),
                    売上金額=("売上金額", "sum"),
                )
                .reset_index()
            )
            # 集計結果を辞書のリスト形式に変換（Excelに書き込むために）
            aggregated_sales = aggregated_df.to_dict("records")
            # 全商品の売上金額の合計を計算
            total_sales = aggregated_df["売上金額"].sum()

        # --- 手数料と支払金額の計算 ---
        # 顧客データから手数料率を取得（文字列形式、例："10%" または "0.1"）
        commission_rate_str = str(customer_data.get("手数料率", "0"))
        # 手数料率を数値に変換
        # "%"が含まれている場合は、それを取り除いて100で割る（例："10%" → 0.1）
        # 含まれていない場合は、そのまま数値に変換（例："0.1" → 0.1）
        commission_rate = (
            float(commission_rate_str.strip("%")) / 100
            if "%" in commission_rate_str
            else float(commission_rate_str)
        )

        # 委託販売手数料 = 売上金額 × 手数料率
        commission_fee = int(total_sales * commission_rate)
        # 消費税 = (売上金額 - 手数料) × 10%
        commission_tax = int((total_sales - commission_fee) * 0.1)
        # お支払金額 = 売上金額 - 手数料 + 消費税 - 振込手数料
        payment_amount = total_sales - commission_fee + commission_tax - bank_transfer_fee

        # 計算したデータをExcelシートに書き込む
        # _fill_settlement_data()は、このクラス内の別のメソッド（内部メソッド）
        # 名前が"_"で始まっているのは、「外部から直接呼ばないでね」という意味
        self._fill_settlement_data(
            ws,
            customer_data,
            aggregated_sales,  # 集計済みのデータを渡す
            period_start,
            period_end,
            issue_date,
            due_date,
            settlement_number,
            total_sales,
            commission_fee,
            commission_tax,
            bank_transfer_fee,
            payment_amount,
        )

    def _fill_settlement_data(
        self,
        ws: Worksheet,
        customer_data,
        sales_details,
        period_start,
        period_end,
        issue_date,
        due_date,
        settlement_number,
        total_sales,
        commission_fee,
        commission_tax,
        bank_transfer_fee,
        payment_amount,
    ):
        """
        精算書のExcelシートに、計算済みのデータを書き込む内部メソッド

        Args:
            ws: データを書き込むExcelシート
            customer_data: 顧客情報の辞書
            sales_details: 集計済みの売上明細のリスト
            period_start: 精算期間の開始日
            period_end: 精算期間の終了日
            issue_date: 発行日
            due_date: 振込予定日
            settlement_number: 精算番号
            total_sales: 売上合計金額
            commission_fee: 委託販売手数料
            commission_tax: 消費税
            bank_transfer_fee: 振込手数料
            payment_amount: お支払金額
        """
        # --- ヘッダー情報の書き込み ---
        # ws["E4"]: E列4行目のセルに値を設定（Excelのセル参照と同じ）
        ws["E4"] = settlement_number  # 精算番号
        # strftime(): 日付を指定した形式の文字列に変換
        # '%Y年%m月%d日' → "2025年10月25日" のような形式
        ws["E5"] = issue_date.strftime("%Y年%m月%d日")  # 発行日

        # --- 顧客情報の書き込み ---
        ws["A10"] = f"{customer_data.get('会社名', '')} 様"  # 会社名
        # 郵便番号と住所を結合して書き込み
        ws["A11"] = f"〒{customer_data.get('郵便番号', '')} {customer_data.get('住所', '')}"
        ws["C10"] = payment_amount  # お支払金額（右上に表示）

        # 精算期間を書き込み
        ws["A15"] = (
            f"精算期間: {period_start.strftime('%Y年%m月%d日')} ～ {period_end.strftime('%Y年%m月%d日')}"
        )

        # --- 売上明細の書き込み ---
        # 売上明細を書き始める行番号（17行目から）
        start_row = 17
        # enumerate(): リストの要素とインデックス（番号）を同時に取得
        # 例: [(0, 商品1), (1, 商品2), ...]
        for idx, sale in enumerate(sales_details):
            # 現在書き込む行番号を計算（17行目、18行目、19行目...）
            row = start_row + idx
            # シートの最大行数を超える場合は、新しい行を挿入
            if row > ws.max_row:
                ws.insert_rows(row)
            # 各列に商品情報を書き込み
            # f"A{row}": f文字列を使って動的にセル参照を作成（例: "A17", "A18"）
            ws[f"A{row}"] = sale.get("商品コード", "")  # A列: 商品コード
            ws[f"B{row}"] = sale.get("商品名", "")  # B列: 商品名
            ws[f"C{row}"] = sale.get("単価", "")  # C列: 単価
            ws[f"D{row}"] = sale.get("販売数", "")  # D列: 販売数
            ws[f"E{row}"] = sale.get("売上金額", "")  # E列: 売上金額

        # --- 計算結果の書き込み ---
        # C27に小計を書き込む
        ws["C27"] = total_sales  # 売上合計金額

        # --- 手数料の書き込み ---
        # 手数料率を表示用の文字列に変換（"10%" の形式にする）
        commission_rate_display = customer_data.get("手数料率", "0%")
        # 文字列でない、または"%"が含まれていない場合は、"%"を追加
        if not isinstance(commission_rate_display, str) or "%" not in commission_rate_display:
            commission_rate_display = f"{commission_rate_display}%"

        # 委託販売手数料
        ws["A28"] = f"委託販売手数料 ({commission_rate_display})"  # A28手数料のラベル
        ws["E28"] = -commission_fee  # マイナス値で表示（引き算の意味）

        # 消費税額
        ws["E29"] = commission_tax  # 消費税額

        # 振込手数料
        ws["E30"] = -bank_transfer_fee  # マイナス値で表示

        # 支払金額
        ws["E31"] = payment_amount  # 最終的な支払金額

        # --- 振込情報の書き込み ---
        # A33に振込情報を書き込む
        ws["A33"] = f"お振込予定日: {due_date.strftime('%Y年%m月%d日')}"

        # 銀行口座情報を1つの文字列にまとめる
        bank_info = (
            f"{customer_data.get('銀行名', '')} "
            f"{customer_data.get('支店名', '')}"
            f"({customer_data.get('支店番号', '')}) "
            f"{customer_data.get('口座種別', '')} "
            f"{customer_data.get('口座番号', '')}"
        )
        ws["A35"] = bank_info  # 銀行口座情報
        ws["A36"] = f"口座名義: {customer_data.get('口座名義', '')}"  # 口座名義


def create_settlements_for_month(
    year: int, month: int, customer_file: str, sales_file: str, template_file: str, output_dir: str
):
    """
    指定された年月の精算書を全顧客分作成し、単一のExcelファイルに出力する関数

    この関数は、以下の処理を行います：
    1. 顧客データと売上データを読み込む
    2. 指定された年月の売上データを抽出
    3. 顧客ごとに精算書シートを作成
    4. 1つのExcelファイルにまとめて保存

    Args:
        year: 精算対象の年（例: 2025）
        month: 精算対象の月（例: 10）
        customer_file: 顧客データが入ったExcelファイルのパス
        sales_file: 売上データが入ったExcelファイルのパス
        template_file: 精算書のテンプレートExcelファイルのパス
        output_dir: 生成された精算書を保存するディレクトリのパス

    Returns:
        str | None: 生成されたExcelファイルのパス。失敗した場合はNone
    """
    print(f"{year}年{month}月分の精算書作成処理を開始します。")

    # --- ファイルの読み込み ---
    try:
        # pandasでExcelファイルを読み込む
        # pd.read_excel(): Excelファイルを読み込んで、DataFrame（表形式のデータ）に変換
        customers_df = pd.read_excel(customer_file)  # 顧客データ
        sales_df = pd.read_excel(sales_file)  # 売上データ

        # openpyxlでテンプレートExcelファイルを読み込む
        # load_workbook(): Excelファイルを読み込んで、Workbookオブジェクトに変換
        main_wb = load_workbook(template_file)
        # .active: アクティブな（最初の）シートを取得
        template_sheet = main_wb.active
    except FileNotFoundError as e:
        # ファイルが見つからない場合のエラー処理
        print(f"エラー: データファイルまたはテンプレートファイルが見つかりません。 {e}")
        # ValueErrorを発生させて、呼び出し元にエラーを伝える
        raise ValueError(f"データファイルまたはテンプレートファイルが見つかりません: {e}")

    # --- 売上データの前処理 ---
    # "売上日"列を日付型に変換（文字列や数値で入っている可能性があるため）
    # pd.to_datetime(): 文字列や数値を日付型に変換
    # .dt.date: 時刻部分を削除して、日付のみにする
    sales_df["売上日"] = pd.to_datetime(sales_df["売上日"]).dt.date

    # --- 対象期間の売上データを抽出 ---
    # calendar.monthrange(): 指定した年月の最初の曜日と最終日を取得
    # 例: (0, 31) → 0は月曜日、31はその月の最終日
    _, last_day = calendar.monthrange(year, month)
    # 精算期間の開始日と終了日を設定
    period_start = date(year, month, 1)  # その月の1日
    period_end = date(year, month, last_day)  # その月の最終日

    # 売上データから、指定された期間のデータだけを抽出
    # 条件: 売上日が期間開始日以降 かつ 期間終了日以前
    month_sales_df = sales_df[(sales_df["売上日"] >= period_start) & (sales_df["売上日"] <= period_end)]

    # 対象期間の売上データが存在しない場合
    if month_sales_df.empty:
        print("警告: 対象期間の売上データがありません。")
        return None  # Noneを返して処理を終了

    # --- 精算書生成の準備 ---
    # SettlementGeneratorクラスのインスタンス（実体）を作成
    # このインスタンスを使って、各顧客の精算書シートを作成します
    generator = SettlementGenerator()

    # 作成したシートの数をカウントする変数
    sheets_created_count = 0

    # --- 顧客ごとに精算書シートを作成 ---
    # customers_df.iterrows(): 顧客データの各行を順番に処理
    # _: 行番号（今回は使わないので、_で無視）
    # customer: その行の顧客データ（pandasのSeriesオブジェクト）
    for _, customer in customers_df.iterrows():
        # 顧客IDと会社名を取得
        client_id = customer["クライアントID"]
        client_name = customer["会社名"]

        # この顧客の売上データだけを抽出
        # 条件: クライアントIDが一致する行だけを取得
        client_sales_df = month_sales_df[month_sales_df["クライアントID"] == client_id]

        # この顧客の売上データが存在しない場合
        if client_sales_df.empty:
            print(f"- {client_name}: 売上データがないためスキップします。")
            continue  # 次の顧客の処理に進む

        print(f"- {client_name}: 精算書シートを作成中...")

        # 顧客データを辞書形式に変換（関数に渡すために）
        customer_data_dict = customer.to_dict()
        # 売上データを辞書のリスト形式に変換（関数に渡すために）
        sales_data_list = client_sales_df.to_dict("records")

        # 精算書シートを作成
        # generator.create_settlement_sheet(): この顧客用の精算書シートを追加
        generator.create_settlement_sheet(
            wb=main_wb,  # シートを追加するExcelブック
            template_sheet=template_sheet,  # コピー元のテンプレートシート
            customer_data=customer_data_dict,  # 顧客情報
            sales_data=sales_data_list,  # 売上明細
            year=year,  # 年
            month=month,  # 月
        )
        # シートを作成したので、カウントを増やす
        sheets_created_count += 1

    # --- ファイルの保存 ---
    if sheets_created_count > 0:
        # 少なくとも1つのシートが作成された場合
        # 元のテンプレートシートを削除（精算書シートだけを残す）
        main_wb.remove(template_sheet)

        # 出力ディレクトリが存在しない場合は作成
        # exist_ok=True: 既に存在していてもエラーにしない
        os.makedirs(output_dir, exist_ok=True)

        # 出力ファイル名を生成
        # f文字列: 変数を文字列に埋め込む（例: "精算書_202510_全顧客.xlsx"）
        # {month:02d}: 月を2桁の数値で表示（例: 10 → "10", 5 → "05"）
        output_filename = os.path.join(output_dir, f"精算書_{year}{month:02d}_全顧客.xlsx")

        # Excelファイルを保存
        main_wb.save(output_filename)
        print(f"\n処理が完了しました。'{output_filename}' に全顧客の精算書が保存されました。")
        # 生成されたファイルのパスを返す
        return output_filename
    else:
        # シートが1つも作成されなかった場合
        print("\n処理対象の顧客がいなかったため、ファイルは作成されませんでした。")
        return None
