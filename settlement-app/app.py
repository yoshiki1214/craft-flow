# -*- coding: utf-8 -*-
"""
委託販売精算書生成アプリケーションのメインファイル
Flaskを使用してWebアプリケーションを構築しています。
"""

# --- モジュールのインポート ---
# Flask: Webフレームワーク
# request: クライアントからのリクエスト情報（フォームデータやファイルなど）を取得
# redirect: 別のURLにリダイレクト
# render_template: HTMLテンプレートを読み込んで表示
# send_from_directory: ディレクトリからファイルを安全に送信
# flash: ユーザーに一時的なメッセージを表示（例：「アップロードが完了しました」）
from flask import Flask, request, redirect, render_template, send_from_directory, flash
import os  # osモジュール: ファイルパスの操作やディレクトリの作成など、OS関連の機能を提供
from werkzeug.utils import secure_filename  # ファイル名を安全な形式に変換（悪意のあるファイル名を無害化）
import pandas as pd  # pandas: データ分析ライブラリ。Excelファイルの読み書きやデータ操作に利用
from datetime import datetime  # datetimeモジュールからdatetimeクラスをインポート: 日付を扱うために使用
from flask_sqlalchemy import SQLAlchemy  # Flask-SQLAlchemy: データベース操作
from flask_migrate import Migrate  # Flask-Migrate: データベースマイグレーション
from services.settlement_generator import (
    create_settlements_for_month,
)  # 自作の精算書生成モジュール

# --- Flaskアプリケーションの初期設定 ---

# Flaskアプリケーションのインスタンスを作成
# __name__ は、このファイルが実行されているモジュール名を示す特殊な変数
# Flaskはこれを使って、テンプレートファイルや静的ファイルを探す場所を特定します
app = Flask(__name__)

# flashメッセージ機能を使うための「秘密鍵」を設定
# このキーはセッション情報を暗号化するために使われます
# 本番環境では、推測されにくい安全な文字列を環境変数から読み込むのが一般的です
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-key-change-in-production")

# データベース設定
base_dir = os.path.dirname(os.path.abspath(__file__))
instance_path = os.path.join(base_dir, "instance")
os.makedirs(instance_path, exist_ok=True)
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL", f"sqlite:///{os.path.join(instance_path, 'settlement.db')}"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# データベース初期化
db = SQLAlchemy(app)
migrate = Migrate(app, db)

# ファイルアップロード先のディレクトリパスを設定
# os.path.abspath(__file__): このファイルの絶対パスを取得 (例: /path/to/your/project/app.py)
# os.path.dirname(...): パスのディレクトリ部分を取得 (例: /path/to/your/project)
# os.path.join(...): OSに合わせてパスを結合 (例: /path/to/your/project/uploads)
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
# Flaskアプリケーションの設定情報として、アップロードフォルダのパスを保存
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# アップロードを許可するファイルの拡張子を定義
ALLOWED_EXTENSIONS = {"xlsx", "xlsm"}


# --- 関数定義 ---


def allowed_file(filename):
    """
    アップロードされたファイル名が、許可された拡張子を持っているかチェックする関数

    Args:
        filename (str): チェックするファイル名

    Returns:
        bool: 許可された拡張子であればTrue、そうでなければFalse
    """
    # 1. ファイル名に"."が含まれているか (拡張子があるか)
    # 2. ファイル名を"."で右から分割し、最後の要素（拡張子）を取得
    # 3. 拡張子を小文字に変換し、許可リスト(ALLOWED_EXTENSIONS)に含まれているか確認
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def _convert_excel_to_pdf(excel_path: str) -> str | None:
    """
    ExcelファイルをPDFに変換する内部関数（openpyxl + reportlab使用）
    Excelのレイアウト、フォント、色、罫線などを可能な限り再現

    Args:
        excel_path: Excelファイルのパス

    Returns:
        str | None: 生成されたPDFファイルのパス。失敗した場合はNone
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.platypus import (
            Table,
            TableStyle,
            SimpleDocTemplate,
            Paragraph,
            Spacer,
            PageBreak,
            KeepTogether,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY

        # PDFファイルのパスを生成（拡張子を.pdfに変更）
        pdf_path = os.path.splitext(excel_path)[0] + ".pdf"

        # Excelファイルを読み込む（data_only=Trueで数式の計算結果を取得）
        wb = load_workbook(excel_path, data_only=True)

        # PDFドキュメントを作成（A4横または縦）
        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=A4,
            rightMargin=15 * mm,
            leftMargin=15 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm,
        )

        story = []
        styles = getSampleStyleSheet()

        # 各シートを処理
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]

            # シート名をタイトルとして追加
            if sheet_idx > 0:
                story.append(PageBreak())

            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=16,
                textColor=colors.HexColor("#000000"),
                spaceAfter=12,
                alignment=TA_CENTER,
            )
            title = Paragraph(f"<b>{sheet_name}</b>", title_style)
            story.append(title)
            story.append(Spacer(1, 12))

            # シートの使用範囲を取得
            if ws.max_row == 0 or ws.max_column == 0:
                continue

            # 列幅を取得（Excelの列幅をmmに変換）
            col_widths = []
            for col_idx in range(1, min(ws.max_column + 1, 15)):  # 最大15列まで
                col_letter = get_column_letter(col_idx)
                # Excelの列幅を取得（デフォルトは約8.43文字分）
                column_dimension = ws.column_dimensions.get(col_letter)
                if column_dimension and column_dimension.width:
                    # Excelの列幅（文字数）をmmに変換（1文字 ≈ 7mm）
                    width_mm = column_dimension.width * 7
                else:
                    width_mm = 8.43 * 7  # デフォルト幅
                col_widths.append(width_mm)

            # データを読み込む（セルの書式、フォント、色などを保持）
            data = []
            row_heights = []

            for row_idx in range(1, min(ws.max_row + 1, 100)):  # 最大100行まで
                row_data = []
                row_height = None

                for col_idx in range(1, min(ws.max_column + 1, 15)):  # 最大15列まで
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value

                    # 行の高さを取得
                    if row_height is None:
                        row_dimension = ws.row_dimensions.get(row_idx)
                        if row_dimension and row_dimension.height:
                            row_height = row_dimension.height * 0.75  # Excelのポイントをmmに変換
                        else:
                            row_height = 15  # デフォルト高さ

                    # 値の処理
                    cell_value = ""
                    if value is None:
                        cell_value = ""
                    elif isinstance(value, datetime):
                        cell_value = value.strftime("%Y年%m月%d日")
                    elif isinstance(value, (int, float)):
                        # 数値のフォーマットを適用
                        if cell.number_format:
                            fmt = str(cell.number_format)
                            if "%" in fmt:
                                cell_value = f"{value:.2f}%"
                            elif "¥" in fmt or "yen" in fmt.lower() or "currency" in fmt.lower():
                                cell_value = f"¥{value:,.0f}"
                            elif "0" in fmt and "," in fmt:
                                cell_value = f"{value:,.0f}"
                            elif "." in fmt:
                                cell_value = f"{value:.2f}"
                            else:
                                cell_value = f"{value:,.0f}" if isinstance(value, float) else str(int(value))
                        else:
                            cell_value = f"{value:,.0f}" if isinstance(value, float) else str(int(value))
                    else:
                        cell_value = str(value)

                    row_data.append(cell_value)

                # 空行でない場合のみ追加
                if any(str(v).strip() for v in row_data):
                    data.append(row_data)
                    row_heights.append(row_height)

            if data:
                # 列幅を調整（ページ幅に収まるように）
                page_width = A4[0] - 30 * mm
                total_width = sum(col_widths) if col_widths else len(data[0]) * 50 * mm
                if total_width > page_width:
                    scale_factor = page_width / total_width
                    col_widths = [w * scale_factor for w in col_widths]

                # テーブルを作成
                table = Table(data, colWidths=col_widths[: len(data[0])] if col_widths else None)

                # テーブルスタイルを構築（Excelの書式を再現）
                table_style = []

                # 各セルのスタイルを適用
                for row_idx, row_data in enumerate(data):
                    for col_idx in range(len(row_data)):
                        cell = ws.cell(row=row_idx + 1, column=col_idx + 1)

                        # フォントスタイル
                        font = cell.font
                        font_name = "Helvetica"
                        font_size = 9
                        font_color = colors.black

                        if font:
                            if font.bold:
                                font_name = "Helvetica-Bold"
                            if font.size:
                                font_size = int(font.size)
                            if font.color and font.color.rgb:
                                try:
                                    rgb = font.color.rgb
                                    if isinstance(rgb, str) and rgb.startswith("FF"):
                                        rgb = rgb[2:]  # 'FF'プレフィックスを削除
                                    font_color = colors.HexColor("#" + rgb)
                                except Exception:
                                    pass

                        # 背景色
                        fill = cell.fill
                        bg_color = None
                        if fill and fill.start_color and fill.start_color.rgb:
                            try:
                                rgb = fill.start_color.rgb
                                if isinstance(rgb, str) and rgb.startswith("FF"):
                                    rgb = rgb[2:]
                                bg_color = colors.HexColor("#" + rgb)
                            except Exception:
                                pass

                        # 配置
                        alignment = cell.alignment
                        align = TA_LEFT
                        if alignment:
                            if alignment.horizontal == "center":
                                align = TA_CENTER
                            elif alignment.horizontal == "right":
                                align = TA_RIGHT
                            elif alignment.horizontal == "justify":
                                align = TA_JUSTIFY

                        # スタイルを適用
                        cell_range = (col_idx, row_idx)
                        table_style.append(("FONTNAME", cell_range, cell_range, font_name))
                        table_style.append(("FONTSIZE", cell_range, cell_range, font_size))
                        table_style.append(("TEXTCOLOR", cell_range, cell_range, font_color))
                        if bg_color:
                            table_style.append(("BACKGROUND", cell_range, cell_range, bg_color))
                        table_style.append(("ALIGN", cell_range, cell_range, align))

                # 罫線を追加
                table_style.append(("GRID", (0, 0), (-1, -1), 0.5, colors.grey))

                # 行の高さを設定
                for row_idx, height in enumerate(row_heights):
                    if height:
                        table_style.append(("ROWBACKGROUNDS", (0, row_idx), (-1, row_idx), [colors.white]))

                table.setStyle(TableStyle(table_style))
                story.append(KeepTogether(table))
                story.append(Spacer(1, 20))

        # PDFを生成
        doc.build(story)

        if os.path.exists(pdf_path):
            return pdf_path
        else:
            return None

    except ImportError as e:
        print(f"必要なライブラリがインストールされていません: {str(e)}")
        print("pip install reportlab Pillow を実行してください")
        return None
    except Exception as e:
        print(f"PDF変換中にエラーが発生しました: {str(e)}")
        import traceback

        traceback.print_exc()
        return None


def _extract_year_month_from_sales(sales_filename: str) -> tuple[int | None, int | None]:
    """
    売上データファイル（Excel）から年と月を自動で抽出する内部関数

    Args:
        sales_filename (str): 売上データのファイル名

    Returns:
        tuple[int | None, int | None]: (年, 月) のタプル。抽出に失敗した場合は (None, None)
    """
    try:
        # このファイルがあるディレクトリの絶対パスを取得
        base_dir = os.path.dirname(os.path.abspath(__file__))
        # 売上データのフルパスを作成
        sales_file = os.path.join(base_dir, "uploads", "sales", sales_filename)

        # ファイルが存在しない場合は、Noneを返す
        if not os.path.exists(sales_file):
            return None, None

        # pandasでExcelファイルを読み込む
        sales_df = pd.read_excel(sales_file)

        # "売上日"という列が存在しない場合は、Noneを返す
        if "売上日" not in sales_df.columns:
            return None, None

        # "売上日"の列を日付型に変換（時刻は切り捨てる）
        sales_df["売上日"] = pd.to_datetime(sales_df["売上日"]).dt.date

        # "売上日"列が空の場合は、Noneを返す
        if sales_df["売上日"].empty:
            return None, None

        # 最も古い日付と最も新しい日付を取得
        min_date = sales_df["売上日"].min()
        max_date = sales_df["売上日"].max()

        # 全てのデータが同じ月のものであれば、その年月を返す
        if min_date.year == max_date.year and min_date.month == max_date.month:
            return min_date.year, min_date.month

        # 複数月にまたがるデータの場合は、最も新しい月を対象とする
        return max_date.year, max_date.month

    except Exception:
        # 何らかのエラーが発生した場合（例：ファイル形式が不正など）はNoneを返す
        return None, None


# モデルの定義（循環インポートを避けるため、ここで定義）
class SettlementHistory(db.Model):
    """
    精算書発行履歴テーブル

    精算書の一括生成・ダウンロード履歴を保存する。

    Attributes:
        id: 主キー（自動採番）
        year: 精算対象年
        month: 精算対象月
        file_name: 生成されたファイル名
        file_path: ファイルの保存パス
        file_format: ファイル形式（excel/pdf）
        created_at: 発行日時（自動設定）
    """

    __tablename__ = "settlement_history"

    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    year = db.Column(db.Integer, nullable=False, comment="精算対象年")
    month = db.Column(db.Integer, nullable=False, comment="精算対象月")
    file_name = db.Column(db.String(255), nullable=False, comment="ファイル名")
    file_path = db.Column(db.String(500), nullable=False, comment="ファイル保存パス")
    file_format = db.Column(db.String(10), nullable=False, default="excel", comment="ファイル形式")
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, comment="発行日時")

    # インデックスの定義
    __table_args__ = (
        db.Index("idx_settlement_history_year_month", "year", "month"),
        db.Index("idx_settlement_history_created_at", "created_at"),
    )

    def __repr__(self) -> str:
        """オブジェクトの文字列表現を返す"""
        return f"<SettlementHistory {self.id}: {self.year}年{self.month}月 - {self.file_name}>"

    def to_dict(self) -> dict:
        """辞書形式に変換"""
        return {
            "id": self.id,
            "year": self.year,
            "month": self.month,
            "file_name": self.file_name,
            "file_path": self.file_path,
            "file_format": self.file_format,
            "created_at": self.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        }


# --- ルート（URL）定義 ---


@app.route("/", methods=["GET"])
def index():
    """
    トップページ（/）にアクセスされたときに呼ばれる関数
    ファイルアップロード用のHTML画面を表示します。
    """
    # "templates"フォルダの中の"upload.html"を読み込んでユーザーに表示
    return render_template("upload.html")


@app.route("/history", methods=["GET"])
def history_page():
    """
    発行履歴一覧ページを表示
    """
    # 発行履歴を新しい順に取得
    histories = SettlementHistory.query.order_by(SettlementHistory.created_at.desc()).all()

    return render_template("history.html", histories=histories)


@app.route("/download/<int:history_id>")
def download_page(history_id):
    """
    精算書ダウンロードページを表示
    """
    history = SettlementHistory.query.get_or_404(history_id)

    # ファイルの存在確認
    if not os.path.exists(history.file_path):
        flash(f"ファイルが見つかりません: {history.file_name}", "error")
        return redirect("/")

    return render_template("download.html", history=history)


@app.route("/download/<int:history_id>/file")
def download_file(history_id):
    """
    精算書ファイルをダウンロード（形式を指定）
    """
    history = SettlementHistory.query.get_or_404(history_id)
    file_format = request.args.get("format", "excel")  # excel または pdf

    # ファイルの存在確認
    if not os.path.exists(history.file_path):
        flash(f"ファイルが見つかりません: {history.file_name}", "error")
        return redirect("/")

    if file_format == "excel":
        # Excelファイルをそのままダウンロード
        return send_from_directory(
            os.path.dirname(history.file_path), os.path.basename(history.file_path), as_attachment=True
        )
    elif file_format == "pdf":
        # PDFに変換してダウンロード
        try:
            pdf_path = _convert_excel_to_pdf(history.file_path)
            if pdf_path and os.path.exists(pdf_path):
                return send_from_directory(
                    os.path.dirname(pdf_path), os.path.basename(pdf_path), as_attachment=True
                )
            else:
                flash("PDF変換に失敗しました。LibreOfficeがインストールされている必要があります。", "error")
                return redirect(f"/download/{history_id}")
        except Exception as e:
            flash(f"PDF変換中にエラーが発生しました: {str(e)}", "error")
            return redirect(f"/download/{history_id}")
    else:
        flash("無効なファイル形式が指定されました", "error")
        return redirect(f"/download/{history_id}")


@app.route("/history/<int:history_id>/download")
def download_history(history_id):
    """
    履歴から過去の精算書をダウンロード（形式を指定）
    """
    history = SettlementHistory.query.get_or_404(history_id)
    file_format = request.args.get("format", "excel")  # excel または pdf

    # ファイルの存在確認
    if not os.path.exists(history.file_path):
        flash(f"ファイルが見つかりません: {history.file_name}", "error")
        return redirect("/history")

    if file_format == "excel":
        # Excelファイルをそのままダウンロード
        return send_from_directory(
            os.path.dirname(history.file_path), os.path.basename(history.file_path), as_attachment=True
        )
    elif file_format == "pdf":
        # PDFに変換してダウンロード
        try:
            pdf_path = _convert_excel_to_pdf(history.file_path)
            if pdf_path and os.path.exists(pdf_path):
                return send_from_directory(
                    os.path.dirname(pdf_path), os.path.basename(pdf_path), as_attachment=True
                )
            else:
                flash("PDF変換に失敗しました。LibreOfficeがインストールされている必要があります。", "error")
                return redirect("/history")
        except Exception as e:
            flash(f"PDF変換中にエラーが発生しました: {str(e)}", "error")
            return redirect("/history")
    else:
        flash("無効なファイル形式が指定されました", "error")
        return redirect("/history")


@app.route("/history/<int:history_id>/delete", methods=["POST"])
def delete_history(history_id):
    """
    発行履歴を削除する
    """
    history = SettlementHistory.query.get_or_404(history_id)

    try:
        # ファイルが存在する場合は削除
        if os.path.exists(history.file_path):
            try:
                os.remove(history.file_path)
                # PDFファイルも存在する場合は削除
                pdf_path = os.path.splitext(history.file_path)[0] + ".pdf"
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
            except Exception as e:
                print(f"ファイル削除エラー: {str(e)}")
                # ファイル削除に失敗してもデータベースレコードは削除する

        # データベースから削除
        file_name = history.file_name
        db.session.delete(history)
        db.session.commit()

        flash(f"発行履歴「{file_name}」を削除しました。", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"削除に失敗しました: {str(e)}", "error")

    return redirect("/history")


@app.route("/upload", methods=["POST"])
def upload_files():
    """
    ファイルアップロード処理のメインとなる関数
    顧客データと売上データの両方がアップロードされたら、自動で精算書を生成します。
    """
    customer_filename = None  # アップロードされた顧客データファイル名を保存する変数
    sales_filename = None  # アップロードされた売上データファイル名を保存する変数

    # --- 顧客データの処理 ---
    # "customer_file"という名前のファイルがフォームから送信されたかチェック
    if "customer_file" in request.files:
        customer_file = request.files["customer_file"]
        # ファイル名が空でないことを確認
        if customer_file.filename != "":
            # ファイルが存在し、かつ許可された拡張子かチェック
            if customer_file and allowed_file(customer_file.filename):
                try:
                    # 保存先フォルダのパスを作成
                    customers_folder = os.path.join(app.config["UPLOAD_FOLDER"], "customers")
                    # フォルダがなければ作成
                    if not os.path.exists(customers_folder):
                        os.makedirs(customers_folder)
                    # ファイル名を安全な形式に変換
                    filename = secure_filename(customer_file.filename)
                    # ファイルを保存
                    customer_file.save(os.path.join(customers_folder, filename))
                    customer_filename = filename  # ファイル名を後で使うために保存
                    # 成功メッセージをflashで表示
                    flash(f"顧客データ: {filename} がアップロードされました", "success")
                except Exception as e:
                    # エラーメッセージをflashで表示
                    flash(
                        f"顧客データ: {customer_file.filename} のアップロードに失敗しました: {str(e)}",
                        "error",
                    )
            else:
                # 許可されていないファイル形式の場合のエラーメッセージ
                flash(
                    f"顧客データ: {customer_file.filename} は許可されていないファイル形式です（Excelファイルのみ許可）",
                    "error",
                )

    # --- 委託販売売上データの処理 (顧客データと同様の流れ) ---
    if "sales_file" in request.files:
        sales_file = request.files["sales_file"]
        if sales_file.filename != "":
            if sales_file and allowed_file(sales_file.filename):
                try:
                    sales_folder = os.path.join(app.config["UPLOAD_FOLDER"], "sales")
                    if not os.path.exists(sales_folder):
                        os.makedirs(sales_folder)
                    filename = secure_filename(sales_file.filename)
                    sales_file.save(os.path.join(sales_folder, filename))
                    sales_filename = filename
                    flash(f"委託販売売上データ: {filename} がアップロードされました", "success")
                except Exception as e:
                    flash(
                        f"委託販売売上データ: {sales_file.filename} のアップロードに失敗しました: {str(e)}",
                        "error",
                    )
            else:
                flash(
                    f"委託販売売上データ: {sales_file.filename} は許可されていないファイル形式です（Excelファイルのみ許可）",
                    "error",
                )

    # --- ファイル選択のバリデーション ---
    customer_selected = "customer_file" in request.files and request.files["customer_file"].filename != ""
    sales_selected = "sales_file" in request.files and request.files["sales_file"].filename != ""
    # どちらのファイルも選択されずにフォームが送信された場合
    if not customer_selected and not sales_selected:
        flash("少なくとも1つのファイルを選択してください", "error")
        return redirect("/")  # トップページに戻る

    # --- 精算書の自動生成処理 ---
    # 顧客データと売上データの両方が正常にアップロードされた場合に実行
    if customer_filename and sales_filename:
        try:
            # 売上データファイルから自動で年月を抽出
            year, month = _extract_year_month_from_sales(sales_filename)

            # 年月が抽出できなかった場合
            if not year or not month:
                flash(
                    "売上データから年と月を抽出できませんでした。売上データに有効な日付が含まれている必要があります。",
                    "error",
                )
                return redirect("/")

            # --- 精算書生成に必要なファイルパスを準備 ---
            base_dir = os.path.dirname(os.path.abspath(__file__))
            customer_file = os.path.join(base_dir, "uploads", "customers", customer_filename)
            sales_file = os.path.join(base_dir, "uploads", "sales", sales_filename)
            template_file = os.path.join(base_dir, "settlement_template.xlsx")
            output_dir = os.path.join(base_dir, "outputs")
            os.makedirs(output_dir, exist_ok=True)  # 出力先フォルダがなければ作成

            # 各ファイルが実際に存在するか確認
            if not os.path.exists(customer_file):
                flash(f"顧客データファイルが見つかりません: {customer_filename}", "error")
                return redirect("/")
            if not os.path.exists(sales_file):
                flash(f"売上データファイルが見つかりません: {sales_filename}", "error")
                return redirect("/")
            if not os.path.exists(template_file):
                flash("テンプレートファイルが見つかりません: settlement_template.xlsx", "error")
                return redirect("/")

            # --- 精算書生成の実行 ---
            # 別のファイル(services/settlement_generator.py)に定義された関数を呼び出す
            output_filename = create_settlements_for_month(
                year=year,
                month=month,
                customer_file=customer_file,
                sales_file=sales_file,
                template_file=template_file,
                output_dir=output_dir,
            )

            # 生成結果に応じてメッセージを表示
            if output_filename:
                output_basename = os.path.basename(output_filename)  # パスからファイル名だけを取得

                # 発行履歴をデータベースに保存
                history = SettlementHistory(
                    year=year,
                    month=month,
                    file_name=output_basename,
                    file_path=output_filename,
                    file_format="excel",
                )
                db.session.add(history)
                db.session.commit()

                # ダウンロードページにリダイレクト
                return redirect(f"/download/{history.id}")
            else:
                flash("精算書の生成に失敗しました。対象期間のデータが存在しない可能性があります。", "error")

        except ValueError as e:
            # 想定される入力値エラー（例：数値変換失敗など）
            flash(f"入力値が不正です: {str(e)}", "error")
        except Exception as e:
            # その他の予期せぬエラー
            flash(f"精算書の自動生成に失敗しました: {str(e)}", "error")

    # 処理が終わったらトップページにリダイレクト
    return redirect("/")


@app.route("/upload/customers", methods=["POST"])
def upload_customers():
    """
    【現在不使用の可能性】顧客データファイル単体をアップロードするための関数
    （メインの/uploadルートで一括処理しているため、通常は使われない）
    """
    if "file" not in request.files:
        flash("ファイルが選択されていません", "error")
        return redirect("/")

    file = request.files["file"]

    if file.filename == "":
        flash("ファイルが選択されていません", "error")
        return redirect("/")

    customers_folder = os.path.join(app.config["UPLOAD_FOLDER"], "customers")
    if not os.path.exists(customers_folder):
        os.makedirs(customers_folder)

    if file and allowed_file(file.filename):
        try:
            filename = secure_filename(file.filename)
            file.save(os.path.join(customers_folder, filename))
            flash(f"顧客データ: {filename} がアップロードされました", "success")
        except Exception:
            flash(f"顧客データ: {file.filename} のアップロードに失敗しました", "error")
    else:
        flash(
            f"顧客データ: {file.filename} は許可されていないファイル形式です（Excelファイルのみ許可）",
            "error",
        )

    return redirect("/")


@app.route("/upload/sales", methods=["POST"])
def upload_sales():
    """
    【現在不使用の可能性】売上データファイル単体をアップロードするための関数
    （メインの/uploadルートで一括処理しているため、通常は使われない）
    """
    if "file" not in request.files:
        flash("ファイルが選択されていません", "error")
        return redirect("/")

    file = request.files["file"]

    if file.filename == "":
        flash("ファイルが選択されていません", "error")
        return redirect("/")

    sales_folder = os.path.join(app.config["UPLOAD_FOLDER"], "sales")
    if not os.path.exists(sales_folder):
        os.makedirs(sales_folder)

    if file and allowed_file(file.filename):
        try:
            filename = secure_filename(file.filename)
            file.save(os.path.join(sales_folder, filename))
            flash(f"委託販売売上データ: {filename} がアップロードされました", "success")
        except Exception:
            flash(f"委託販売売上データ: {file.filename} のアップロードに失敗しました", "error")
    else:
        flash(
            f"委託販売売上データ: {file.filename} は許可されていないファイル形式です（Excelファイルのみ許可）",
            "error",
        )

    return redirect("/")


@app.route("/uploads/<filename>")
def uploaded_file(filename):
    """
    アップロードされたファイルをブラウザで表示・ダウンロードさせるための関数
    （例：/uploads/my_data.xlsx にアクセスするとファイルが返される）
    ※セキュリティ上、直接ファイルシステムにアクセスさせないためにこの関数が存在します
    """
    # send_from_directoryを使うことで、指定されたディレクトリ外のファイルへのアクセスを防ぐ
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


@app.route("/generate", methods=["POST"])
def generate_settlements():
    """
    【現在不使用の可能性】手動で精算書を生成するための関数
    （/uploadルートでの自動生成がメインのため、通常は使われない）
    """
    try:
        # フォームから年、月、ファイル名を取得
        year = int(request.form.get("year"))
        month = int(request.form.get("month"))
        customer_filename = request.form.get("customer_file")
        sales_filename = request.form.get("sales_file")

        # バリデーション（入力値のチェック）
        if not customer_filename or not sales_filename:
            flash("顧客データファイルと売上データファイルの両方を指定してください", "error")
            return redirect("/")
        if not (1 <= month <= 12):
            flash("月は1から12の間で指定してください", "error")
            return redirect("/")

        # ファイルパスの構築
        base_dir = os.path.dirname(os.path.abspath(__file__))
        customer_file = os.path.join(base_dir, "uploads", "customers", customer_filename)
        sales_file = os.path.join(base_dir, "uploads", "sales", sales_filename)
        template_file = os.path.join(
            base_dir, "委託販売精算書.xlsx"
        )  # 注意：テンプレート名がハードコードされている

        # ファイルの存在確認
        if not os.path.exists(customer_file):
            flash(f"顧客データファイルが見つかりません: {customer_filename}", "error")
            return redirect("/")
        if not os.path.exists(sales_file):
            flash(f"売上データファイルが見つかりません: {sales_filename}", "error")
            return redirect("/")
        if not os.path.exists(template_file):
            flash("テンプレートファイルが見つかりません: 委託販売精算書.xlsx", "error")
            return redirect("/")

        # 出力ディレクトリの準備
        output_dir = os.path.join(base_dir, "outputs")
        os.makedirs(output_dir, exist_ok=True)

        # 精算書生成の実行
        output_filename = create_settlements_for_month(
            year=year,
            month=month,
            customer_file=customer_file,
            sales_file=sales_file,
            template_file=template_file,
            output_dir=output_dir,
        )

        # 結果の表示
        if output_filename:
            output_basename = os.path.basename(output_filename)
            flash(f"精算書の生成が完了しました: {output_basename}", "success")
        else:
            flash("精算書の生成に失敗しました。対象期間のデータが存在しない可能性があります。", "error")

    except ValueError as e:
        flash(f"入力値が不正です: {str(e)}", "error")
    except Exception as e:
        flash(f"精算書の生成に失敗しました: {str(e)}", "error")

    return redirect("/")


@app.route("/outputs/<filename>")
def download_output(filename):
    """
    生成された精算書ファイルをダウンロードさせるための関数
    （例：/outputs/2025年10月度_精算書.xlsx にアクセスするとダウンロードが始まる）
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(base_dir, "outputs")
    # send_from_directoryで安全にファイルを送信
    return send_from_directory(output_dir, filename)


# --- アプリケーションの実行 ---

# このファイルが `python app.py` のように直接実行された場合にのみ、以下のコードが実行される
if __name__ == "__main__":
    """
    開発用Webサーバーを起動します。
    """
    # データベーステーブルの作成
    with app.app_context():
        db.create_all()

    # 起動前に、アップロード用のフォルダが存在するか確認し、なければ作成する
    # これにより、初回起動時にアップロードでエラーが出るのを防ぎます
    if not os.path.exists(app.config["UPLOAD_FOLDER"]):
        os.makedirs(app.config["UPLOAD_FOLDER"])

    # Flaskに組み込まれている開発用サーバーを起動
    # debug=True: デバッグモードを有効にする。
    #   - コードを変更するとサーバーが自動で再起動する
    #   - エラーが発生した際に、ブラウザ上で詳細なエラー情報を確認できる
    app.run(debug=True)
