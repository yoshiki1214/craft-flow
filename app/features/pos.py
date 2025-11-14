"""
POS売上データ管理機能

PDFファイルのアップロードとデータベースへの保存を処理する。
"""

import os
import sys
import logging
from io import BytesIO
from typing import Dict, List
from flask import Blueprint, flash, redirect, render_template, request, url_for, send_file
from werkzeug.utils import secure_filename
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app import db
from app.models.pos_sales import PosSales
from app.models.daily_sales import DailySales
from app.utils.pdf_processor import (
    extract_metadata_from_pdf,
    extract_table_data_from_pdf,
    parse_sales_data,
)

pos_bp = Blueprint("pos", __name__, url_prefix="/pos")


def allowed_file(filename: str) -> bool:
    """
    アップロードされたファイルがPDFかどうかをチェックする

    Args:
        filename: ファイル名

    Returns:
        PDFファイルの場合はTrue、それ以外はFalse
    """
    return "." in filename and filename.rsplit(".", 1)[1].lower() == "pdf"


def save_pdf_data_to_db(
    sales_records: List[Dict[str, any]],
    pdf_filename: str,
    overwrite: bool = False,
) -> Dict[str, int]:
    """
    PDFから抽出したデータをpos_salesテーブルに保存する

    Args:
        sales_records: 売上データのリスト
        pdf_filename: PDFファイル名
        overwrite: 上書きオプションが有効かどうか

    Returns:
        処理結果の統計情報（inserted, skipped, overwritten）
    """
    stats = {"inserted": 0, "skipped": 0, "overwritten": 0}

    if not sales_records:
        return stats

    # 最初のレコードからメタデータを取得
    first_record = sales_records[0]
    pos_number = first_record.get("pos_number")
    sale_date = first_record.get("sale_date")
    reported_at = first_record.get("reported_at")

    if not pos_number or not sale_date:
        return stats

    try:
        # 既存データをチェック
        existing_records = PosSales.query.filter_by(pos_number=pos_number, sale_date=sale_date).all()

        if existing_records:
            # 既存データが存在する場合
            if overwrite:
                # 上書きオプションが有効な場合
                existing_reported_at = existing_records[0].reported_at

                # reported_atを比較（文字列として比較）
                if reported_at and reported_at > existing_reported_at:
                    # 新しいデータの方が新しい場合、古いデータを削除
                    for record in existing_records:
                        db.session.delete(record)
                    db.session.commit()

                    # 新しいデータを挿入
                    for record_data in sales_records:
                        pos_sale = PosSales(
                            pos_number=record_data["pos_number"],
                            sale_date=record_data["sale_date"],
                            reported_at=record_data["reported_at"],
                            product_code=record_data["product_code"],
                            product_name=record_data["product_name"],
                            quantity=record_data["quantity"],
                            unit_price=record_data["unit_price"],
                            subtotal=record_data["subtotal"],
                            total_amount=record_data["total_amount"],
                            pdf_source_file=pdf_filename,
                        )
                        db.session.add(pos_sale)
                        stats["overwritten"] += 1

                    db.session.commit()
                else:
                    # 新しいデータの方が古い場合はスキップ
                    stats["skipped"] = len(sales_records)
            else:
                # 上書きオプションが無効な場合はスキップ
                stats["skipped"] = len(sales_records)
        else:
            # 既存データが存在しない場合、新規挿入
            for record_data in sales_records:
                try:
                    pos_sale = PosSales(
                        pos_number=record_data["pos_number"],
                        sale_date=record_data["sale_date"],
                        reported_at=record_data["reported_at"],
                        product_code=record_data["product_code"],
                        product_name=record_data["product_name"],
                        quantity=record_data["quantity"],
                        unit_price=record_data["unit_price"],
                        subtotal=record_data["subtotal"],
                        total_amount=record_data["total_amount"],
                        pdf_source_file=pdf_filename,
                    )
                    db.session.add(pos_sale)
                    stats["inserted"] += 1
                except Exception as e:
                    # UNIQUE制約違反などのエラーを捕捉
                    db.session.rollback()
                    stats["skipped"] += 1
                    print(f"データ挿入エラー: {e}")

            db.session.commit()

    except Exception as e:
        db.session.rollback()
        print(f"DB保存エラー: {e}")
        stats["skipped"] = len(sales_records)

    return stats


def aggregate_daily_sales(sale_date: str = None) -> Dict[str, any]:
    """
    pos_salesテーブルから日次売上を集計し、daily_salesテーブルに保存する

    Args:
        sale_date: 集計対象の営業日（YYYY-MM-DD形式）。Noneの場合は全営業日を集計

    Returns:
        集計結果の統計情報（aggregated_dates: 集計した日付のリスト）
    """
    try:
        # 集計対象の日付を取得
        if sale_date:
            # 特定の日付のみ集計
            dates_to_aggregate = [sale_date]
        else:
            # 全営業日を集計
            dates_to_aggregate = (
                db.session.query(PosSales.sale_date).distinct().order_by(PosSales.sale_date.desc()).all()
            )
            dates_to_aggregate = [date[0] for date in dates_to_aggregate]

        aggregated_dates = []

        for target_date in dates_to_aggregate:
            # 該当日の合計売上を集計
            # Issue #15の要件: SUM(total_amount) AS subtotal FROM pos_sales GROUP BY sale_date
            # ただし、total_amountはPOSレジごとの総合計なので、重複を避けるため
            # 各POSレジのtotal_amountを1回だけカウントする必要がある
            # 実際には、各商品のsubtotalの合計を計算する方が正確
            result = (
                db.session.query(db.func.sum(PosSales.subtotal).label("total_sales"))
                .filter(PosSales.sale_date == target_date)
                .first()
            )

            if result and result.total_sales:
                total_sales_amount = int(result.total_sales)

                # daily_salesテーブルに保存または更新
                existing_daily_sale = DailySales.query.filter_by(sale_date=target_date).first()

                if existing_daily_sale:
                    # 既存データを更新
                    existing_daily_sale.total_sales_amount = total_sales_amount
                    from datetime import datetime

                    existing_daily_sale.created_at = datetime.utcnow()
                else:
                    # 新規データを挿入
                    daily_sale = DailySales(
                        sale_date=target_date,
                        total_sales_amount=total_sales_amount,
                    )
                    db.session.add(daily_sale)

                aggregated_dates.append(target_date)

        db.session.commit()
        print(f"[DEBUG] 日次売上集計完了: {len(aggregated_dates)}件の日付を集計しました")
        sys.stdout.flush()

        return {"aggregated_dates": aggregated_dates, "count": len(aggregated_dates)}

    except Exception as e:
        db.session.rollback()
        import traceback

        print(f"[ERROR] 日次売上集計エラー: {e}")
        print(f"[ERROR] トレースバック:\n{traceback.format_exc()}")
        sys.stdout.flush()
        return {"aggregated_dates": [], "count": 0, "error": str(e)}


@pos_bp.route("/")
def dashboard():
    """
    POSデータ管理ダッシュボード

    Returns:
        ダッシュボードページのHTML
    """
    # 統計情報を取得
    total_records = PosSales.query.count()

    # 営業日とPOSレジ番号の組み合わせを取得（「営業日 POSレジ番号」形式）
    # 重複を排除して、読み込んだファイルの情報を表示
    date_pos_combinations = (
        PosSales.query.with_entities(PosSales.sale_date, PosSales.pos_number)
        .distinct()
        .order_by(PosSales.sale_date.desc(), PosSales.pos_number)
        .all()
    )

    # 「営業日 POSレジ番号」形式のリストを作成
    date_pos_list = []
    from datetime import datetime

    for sale_date, pos_number in date_pos_combinations:
        # 日付を読みやすい形式に変換（例: 2025-11-05 → 2025年11月5日）
        try:
            date_obj = datetime.strptime(str(sale_date), "%Y-%m-%d")
            formatted_date = date_obj.strftime("%Y年%m月%d日")
        except (ValueError, TypeError):
            formatted_date = str(sale_date)

        date_pos_list.append(
            {
                "formatted": f"{formatted_date} {pos_number}",
                "sale_date": str(sale_date),
                "pos_number": str(pos_number),
            }
        )

    return render_template(
        "pos/dashboard.html",
        total_records=total_records,
        date_pos_list=date_pos_list,
    )


@pos_bp.route("/upload", methods=["GET", "POST"])
def upload():
    """
    PDFファイルのアップロード処理

    GET: アップロードフォームを表示
    POST: アップロードされたPDFファイルを処理してDBに保存
    """
    if request.method == "GET":
        return render_template("pos/upload.html")

    # POSTリクエストの処理
    if "files" not in request.files:
        flash("ファイルが選択されていません。", "error")
        return redirect(url_for("pos.upload"))

    files = request.files.getlist("files")
    overwrite = request.form.get("overwrite") == "true"

    if not files or all(f.filename == "" for f in files):
        flash("ファイルが選択されていません。", "error")
        return redirect(url_for("pos.upload"))

    # アップロードディレクトリの作成
    upload_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads")
    os.makedirs(upload_dir, exist_ok=True)

    total_stats = {"inserted": 0, "skipped": 0, "overwritten": 0}
    processed_files = 0
    error_files = []

    for file in files:
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(upload_dir, filename)

            try:
                # ファイルを保存
                file.save(filepath)
                # ログを確実に出力
                logger = logging.getLogger(__name__)
                logger.info(f"[DEBUG] ファイルを保存しました: {filename}")
                print(f"[DEBUG] ファイルを保存しました: {filename}", flush=True)
                sys.stdout.flush()

                # PDFからメタデータを抽出
                metadata = extract_metadata_from_pdf(filepath)
                logger.info(f"[DEBUG] メタデータ抽出結果: {metadata}")
                print(f"[DEBUG] メタデータ抽出結果: {metadata}", flush=True)
                sys.stdout.flush()

                # メタデータの検証
                if not metadata.get("pos_number") or not metadata.get("sale_date"):
                    error_files.append(f"{filename} (メタデータ不足)")
                    pos_num = metadata.get("pos_number")
                    sale_dt = metadata.get("sale_date")
                    print(
                        f"[ERROR] メタデータが不足しています: " f"pos_number={pos_num}, sale_date={sale_dt}"
                    )
                    if os.path.exists(filepath):
                        os.remove(filepath)
                    continue

                # PDFからテーブルデータを抽出
                table_data = extract_table_data_from_pdf(filepath)
                print(f"[DEBUG] テーブルデータ抽出結果: {len(table_data)}行")

                if not table_data:
                    error_files.append(f"{filename} (テーブルデータなし)")
                    print("[ERROR] テーブルデータが抽出できませんでした")
                    if os.path.exists(filepath):
                        os.remove(filepath)
                    continue

                # データを整形
                sales_records = parse_sales_data(table_data, metadata)
                print(f"[DEBUG] 整形後のレコード数: {len(sales_records)}")

                if sales_records:
                    # DBに保存
                    stats = save_pdf_data_to_db(sales_records, filename, overwrite)
                    print(f"[DEBUG] DB保存結果: {stats}")
                    total_stats["inserted"] += stats["inserted"]
                    total_stats["skipped"] += stats["skipped"]
                    total_stats["overwritten"] += stats["overwritten"]
                    processed_files += 1
                else:
                    error_files.append(f"{filename} (有効なレコードなし)")
                    print(
                        f"[ERROR] 有効なレコードが生成されませんでした。テーブルデータ: {len(table_data)}行"
                    )

                # 一時ファイルを削除
                if os.path.exists(filepath):
                    os.remove(filepath)

            except Exception as e:
                error_files.append(f"{filename} ({str(e)})")
                import traceback

                print(f"[ERROR] ファイル処理エラー ({filename}): {e}")
                print(f"[ERROR] トレースバック:\n{traceback.format_exc()}")
                if os.path.exists(filepath):
                    os.remove(filepath)

    # 結果メッセージを生成
    if processed_files > 0:
        message = f"{processed_files}件のファイルを取り込みました。"
        if total_stats["inserted"] > 0:
            message += f" {total_stats['inserted']}件を新規登録しました。"
        if total_stats["skipped"] > 0:
            message += f" {total_stats['skipped']}件は重複のためスキップしました。"
        if total_stats["overwritten"] > 0:
            message += f" {total_stats['overwritten']}件を上書きしました。"
        flash(message, "success")
    else:
        flash("処理できたファイルがありませんでした。", "error")

    if error_files:
        flash(f"エラーが発生したファイル: {', '.join(error_files)}", "warning")

    # アップロード完了後、集計を実行
    if processed_files > 0:
        try:
            # アップロードされたファイルの営業日を取得して集計
            # 最新のデータから営業日を取得
            latest_records = PosSales.query.order_by(PosSales.registration_date.desc()).limit(100).all()
            uploaded_dates = set()
            for record in latest_records:
                uploaded_dates.add(record.sale_date)

            for sale_date in uploaded_dates:
                aggregate_daily_sales(sale_date)

            # 集計完了後、最新の日付の集計結果ページにリダイレクト
            if uploaded_dates:
                latest_date = max(uploaded_dates)
                return redirect(url_for("pos.results", sale_date=latest_date))
        except Exception as e:
            print(f"[WARNING] 集計処理でエラーが発生しました: {e}")
            sys.stdout.flush()

    return redirect(url_for("pos.dashboard"))


@pos_bp.route("/aggregate", methods=["POST"])
def aggregate():
    """
    日次売上集計を実行する

    Returns:
        集計結果ページへのリダイレクト
    """
    try:
        # 全営業日の集計を実行
        result = aggregate_daily_sales()

        if result.get("count", 0) > 0:
            flash(f"{result['count']}件の営業日の集計が完了しました。", "success")
            # 最新の集計結果ページにリダイレクト
            if result.get("aggregated_dates"):
                latest_date = result["aggregated_dates"][0]
                return redirect(url_for("pos.results", sale_date=latest_date))
        else:
            flash("集計対象のデータがありませんでした。", "warning")

    except Exception as e:
        flash(f"集計処理でエラーが発生しました: {str(e)}", "error")

    return redirect(url_for("pos.dashboard"))


@pos_bp.route("/results/<sale_date>")
def results(sale_date: str):
    """
    指定された営業日の集計結果を表示する

    Args:
        sale_date: 営業日（YYYY-MM-DD形式）

    Returns:
        集計結果ページのHTML
    """
    from datetime import datetime

    # daily_salesテーブルから日次サマリを取得
    daily_sale = DailySales.query.filter_by(sale_date=sale_date).first()

    # pos_salesテーブルから商品別集計を取得
    # 商品コード、商品名ごとに数量と合計金額を集計
    product_summary = (
        db.session.query(
            PosSales.product_code,
            PosSales.product_name,
            db.func.sum(PosSales.quantity).label("total_quantity"),
            db.func.sum(PosSales.subtotal).label("total_subtotal"),
        )
        .filter(PosSales.sale_date == sale_date)
        .group_by(PosSales.product_code, PosSales.product_name)
        .order_by(PosSales.product_code)  # 商品コードでソート
        .all()
    )

    # 日付を読みやすい形式に変換
    try:
        date_obj = datetime.strptime(sale_date, "%Y-%m-%d")
        formatted_date = date_obj.strftime("%Y年%m月%d日")
    except (ValueError, TypeError):
        formatted_date = sale_date

    return render_template(
        "pos/results.html",
        sale_date=sale_date,
        formatted_date=formatted_date,
        daily_sale=daily_sale,
        product_summary=product_summary,
    )


@pos_bp.route("/details/<sale_date>/<pos_number>")
def details(sale_date: str, pos_number: str):
    """
    指定された営業日とPOSレジ番号の詳細データを表示する

    Args:
        sale_date: 営業日（YYYY-MM-DD形式）
        pos_number: POSレジ番号

    Returns:
        詳細データページのHTML
    """
    from datetime import datetime

    # pos_salesテーブルから該当するデータを取得
    sales_records = (
        PosSales.query.filter_by(sale_date=sale_date, pos_number=pos_number)
        .order_by(PosSales.product_code)
        .all()
    )

    # 日付を読みやすい形式に変換
    try:
        date_obj = datetime.strptime(sale_date, "%Y-%m-%d")
        formatted_date = date_obj.strftime("%Y年%m月%d日")
    except (ValueError, TypeError):
        formatted_date = sale_date

    # 合計金額を計算
    total_amount = sum(record.subtotal for record in sales_records) if sales_records else 0

    return render_template(
        "pos/details.html",
        sale_date=sale_date,
        formatted_date=formatted_date,
        pos_number=pos_number,
        sales_records=sales_records,
        total_amount=total_amount,
    )


def generate_sales_excel(sale_date: str) -> BytesIO:
    """
    指定された営業日のPOS売上データを集計し、Excelファイルを生成する

    Args:
        sale_date: 営業日（YYYY-MM-DD形式）

    Returns:
        ExcelファイルのBytesIOオブジェクト

    Excelファイルの構成:
    - POS1シート: POSレジ番号「POS1」の商品別集計
    - POS2シート: POSレジ番号「POS2」の商品別集計
    - POS3シート: POSレジ番号「POS3」の商品別集計
    - POS4シート: POSレジ番号「POS4」の商品別集計
    - 売上集計シート: 全レジ（POS1〜4）のデータを合算した商品別集計
    """
    try:
        # pos_salesテーブルから対象日のデータを取得
        sales_data = (
            PosSales.query.filter_by(sale_date=sale_date)
            .with_entities(
                PosSales.pos_number,
                PosSales.product_code,
                PosSales.product_name,
                PosSales.unit_price,
                PosSales.quantity,
                PosSales.subtotal,
            )
            .all()
        )

        if not sales_data:
            raise ValueError(f"{sale_date}のデータが見つかりませんでした")

        # データをpandas DataFrameに変換
        df = pd.DataFrame(
            [
                {
                    "pos_number": record.pos_number,
                    "product_code": record.product_code,
                    "product_name": record.product_name,
                    "unit_price": record.unit_price,
                    "quantity": record.quantity,
                    "subtotal": record.subtotal,
                }
                for record in sales_data
            ]
        )

        # Excelファイルをメモリ上に生成
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # 各POSレジ（POS1〜POS4）のシートを作成
            for pos_num in ["POS1", "POS2", "POS3", "POS4"]:
                pos_df = df[df["pos_number"] == pos_num].copy()

                if not pos_df.empty:
                    # 商品コード、商品名、単価でグループ化して集計
                    aggregated = (
                        pos_df.groupby(["product_code", "product_name", "unit_price"])
                        .agg({"quantity": "sum", "subtotal": "sum"})
                        .reset_index()
                    )
                    aggregated = aggregated.sort_values("product_code")
                else:
                    # データがない場合でも空のDataFrameを作成
                    aggregated = pd.DataFrame(
                        columns=["product_code", "product_name", "unit_price", "quantity", "subtotal"]
                    )

                # カラム名を日本語に変更
                aggregated.columns = ["商品コード", "商品名", "単価", "販売数", "売上金額"]
                aggregated.to_excel(writer, sheet_name=pos_num, index=False)

            # 売上集計シート（全レジ合計）を作成
            aggregated_all = (
                df.groupby(["product_code", "product_name", "unit_price"])
                .agg({"quantity": "sum", "subtotal": "sum"})
                .reset_index()
            )
            aggregated_all = aggregated_all.sort_values("product_code")
            aggregated_all.columns = ["商品コード", "商品名", "単価", "販売数", "売上金額"]
            aggregated_all.to_excel(writer, sheet_name="売上集計", index=False)

            # スタイルの適用（openpyxlを使用）
            # スタイル定義
            header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            row_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
            border_side = Side(style="thin", color="000000")
            border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
            number_format = "#,##0"
            alignment_center = Alignment(horizontal="center", vertical="center")
            alignment_right = Alignment(horizontal="right", vertical="center")
            total_font = Font(bold=True, size=11)

            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]

                # ヘッダー行のスタイル適用
                header_row = worksheet[1]
                for cell in header_row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = alignment_center
                    cell.border = border

                # データ行のスタイル適用と合計計算
                total_quantity = 0
                total_subtotal = 0

                for row_idx, row in enumerate(
                    worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2
                ):
                    # 1行毎に行の背景色を設定（偶数行）
                    if row_idx % 2 == 0:
                        for cell in row:
                            cell.fill = row_fill

                    for idx, cell in enumerate(row):
                        cell.border = border
                        # 列インデックスに応じてスタイルを適用
                        # 0: 商品コード, 1: 商品名, 2: 単価, 3: 販売数, 4: 売上金額
                        if idx == 0:  # 商品コード
                            cell.alignment = alignment_center
                        elif idx == 1:  # 商品名
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        elif idx == 2:  # 単価
                            cell.number_format = number_format
                            cell.alignment = alignment_right
                        elif idx == 3:  # 販売数
                            cell.alignment = alignment_right
                            if cell.value is not None:
                                try:
                                    total_quantity += int(cell.value)
                                except (ValueError, TypeError):
                                    pass
                        elif idx == 4:  # 売上金額
                            cell.number_format = number_format
                            cell.alignment = alignment_right
                            if cell.value is not None:
                                try:
                                    total_subtotal += int(cell.value)
                                except (ValueError, TypeError):
                                    pass

                # 合計行を追加
                total_row_num = worksheet.max_row + 1
                # 合計行のセルを作成
                worksheet.cell(row=total_row_num, column=1, value="合計")
                worksheet.cell(row=total_row_num, column=2, value="")
                worksheet.cell(row=total_row_num, column=3, value="")
                worksheet.cell(row=total_row_num, column=4, value=total_quantity)
                worksheet.cell(row=total_row_num, column=5, value=total_subtotal)

                # 合計行のスタイル適用
                total_row = worksheet[total_row_num]
                for idx, cell in enumerate(total_row):
                    cell.border = border
                    cell.font = total_font
                    if idx == 0:  # 合計
                        cell.alignment = alignment_center
                        if total_row_num % 2 == 0:
                            cell.fill = row_fill
                    elif idx == 1:  # 商品名（空）
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        if total_row_num % 2 == 0:
                            cell.fill = row_fill
                    elif idx == 2:  # 単価（空）
                        cell.alignment = alignment_right
                        if total_row_num % 2 == 0:
                            cell.fill = row_fill
                    elif idx == 3:  # 販売数
                        cell.alignment = alignment_right
                        if total_row_num % 2 == 0:
                            cell.fill = row_fill
                    elif idx == 4:  # 売上金額
                        cell.number_format = number_format
                        cell.alignment = alignment_right
                        if total_row_num % 2 == 0:
                            cell.fill = row_fill

                # 列幅の調整
                column_widths = {
                    "A": 10,  # 商品コード
                    "B": 30,  # 商品名
                    "C": 8,  # 単価
                    "D": 8,  # 販売数
                    "E": 15,  # 売上金額
                }
                for col_letter, width in column_widths.items():
                    worksheet.column_dimensions[col_letter].width = width

                # 行の高さを調整
                worksheet.row_dimensions[1].height = 25  # ヘッダー行
                for row_num in range(2, worksheet.max_row + 1):
                    worksheet.row_dimensions[row_num].height = 20

        output.seek(0)
        return output

    except Exception as e:
        import traceback

        print(f"[ERROR] Excel生成エラー: {e}")
        print(f"[ERROR] トレースバック:\n{traceback.format_exc()}")
        sys.stdout.flush()
        raise


@pos_bp.route("/download/excel/<sale_date>")
def download_excel(sale_date: str):
    """
    Excelファイルを生成してダウンロードする

    Args:
        sale_date: 営業日（YYYY-MM-DD形式）

    Returns:
        Excelファイルのダウンロードレスポンス
    """
    try:
        # Excelファイルを生成
        excel_file = generate_sales_excel(sale_date)

        # ファイル名を生成（例: 商品別売上集計_2025-11-05.xlsx）
        filename = f"商品別売上集計_{sale_date}.xlsx"

        return send_file(
            excel_file,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    except ValueError as e:
        flash(f"エラー: {str(e)}", "error")
        return redirect(url_for("pos.dashboard"))
    except Exception as e:
        flash(f"Excelファイルの生成に失敗しました: {str(e)}", "error")
        return redirect(url_for("pos.dashboard"))


# テスト・デバッグ用: 直接実行時のエラーハンドリング
