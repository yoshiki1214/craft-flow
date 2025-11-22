# -*- coding: utf-8 -*-
"""
委託販売精算書生成アプリケーションのメインファイル
Flaskを使用してWebアプリケーションを構築しています。
"""

# --- モジュールのインポート ---
# Flask: Webフレームワーク
# request: クライアントからのリクエスト情報（フォームデータやファイルなど）を取得
# redirect: 別のURLにリダイレクト
# render_template: HTMLテンプレートを読み込んで表示
# send_from_directory: ディレクトリからファイルを安全に送信
# flash: ユーザーに一時的なメッセージを表示（例：「アップロードが完了しました」）
from flask import Flask, request, redirect, render_template, send_from_directory, flash
import os  # osモジュール: ファイルパスの操作やディレクトリの作成など、OS関連の機能を提供
from werkzeug.utils import secure_filename  # ファイル名を安全な形式に変換（悪意のあるファイル名を無害化）
import pandas as pd  # pandas: データ分析ライブラリ。Excelファイルの読み書きやデータ操作に利用
from datetime import datetime  # datetimeモジュールからdatetimeクラスをインポート: 日付を扱うために使用
from flask_sqlalchemy import SQLAlchemy  # Flask-SQLAlchemy: データベース操作
from flask_migrate import Migrate  # Flask-Migrate: データベースマイグレーション
from services.settlement_generator import (
    create_settlements_for_month,
)  # 自作の精算書生成モジュール

# --- Flaskアプリケーションの初期設定 ---
import sys
import logging

from app import create_app


# ログ設定: 標準出力に確実に出力されるようにする
logging.basicConfig(
    level=logging.DEBUG, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", stream=sys.stdout
)
# 環境変数から設定名を取得
# FLASK_CONFIG環境変数が設定されていない場合は'default'（開発環境）を使用
config_name = os.environ.get("FLASK_CONFIG", "default")
app = create_app()

# Flaskアプリケーションのインスタンスを作成
# __name__ は、このファイルが実行されているモジュール名を示す特殊な変数
# Flaskはこれを使って、テンプレートファイルや静的ファイルを探す場所を特定します
app = Flask(__name__)

# flashメッセージ機能を使うための「秘密鍵」を設定
# このキーはセッション情報を暗号化するために使われます
# 本番環境では、推測されにくい安全な文字列を環境変数から読み込むのが一般的です
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-key-change-in-production")

# データベース設定
base_dir = os.path.dirname(os.path.abspath(__file__))
instance_path = os.path.join(base_dir, "instance")
os.makedirs(instance_path, exist_ok=True)
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL", f"sqlite:///{os.path.join(instance_path, 'settlement.db')}"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# データベース初期化
db = SQLAlchemy(app)
migrate = Migrate(app, db)

# ファイルアップロード先のディレクトリパスを設定
# os.path.abspath(__file__): このファイルの絶対パスを取得 (例: /path/to/your/project/app.py)
# os.path.dirname(...): パスのディレクトリ部分を取得 (例: /path/to/your/project)
# os.path.join(...): OSに合わせてパスを結合 (例: /path/to/your/project/uploads)
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
# Flaskアプリケーションの設定情報として、アップロードフォルダのパスを保存
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# アップロードを許可するファイルの拡張子を定義
ALLOWED_EXTENSIONS = {"xlsx", "xlsm"}


# --- 関数定義 ---


def allowed_file(filename):
    """
    アップロードされたファイル名が、許可された拡張子を持っているかチェックする関数

    Args:
        filename (str): チェックするファイル名

    Returns:
        bool: 許可された拡張子であればTrue、そうでなければFalse
    """
    # 1. ファイル名に"."が含まれているか (拡張子があるか)
    # 2. ファイル名を"."で右から分割し、最後の要素（拡張子）を取得
    # 3. 拡張子を小文字に変換し、許可リスト(ALLOWED_EXTENSIONS)に含まれているか確認
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def _convert_excel_to_pdf(excel_path: str) -> tuple[str | None, str | None]:
    """
    ExcelファイルをPDFに変換する内部関数（LibreOffice使用）
    LibreOfficeのコマンドラインツールを使用してExcelをPDFに変換
    Excelのレイアウトや書式を維持したままPDFに変換できます
    各シートが1ページとして出力されます
    縦向きA4で横幅を1ページに収める設定を適用します
    印刷範囲はA1セルからE39セルまでに限定されます

    Args:
        excel_path: Excelファイルのパス

    Returns:
        tuple[str | None, str | None]: (生成されたPDFファイルのパス, エラーメッセージ)
        成功した場合は (pdf_path, None)、失敗した場合は (None, error_message)
    """
    try:
        import subprocess
        import shutil
        from openpyxl import load_workbook

        # Excelファイルが存在するか確認
        if not os.path.exists(excel_path):
            return None, f"Excelファイルが見つかりません: {excel_path}"

        # 印刷範囲とページ設定を自動設定するため、一時的にExcelファイルを読み込む
        temp_excel_path = None
        try:
            wb = load_workbook(excel_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # 使用されている範囲を取得
                if ws.max_row > 0 and ws.max_column > 0:
                    # 印刷範囲を設定（F39セルまで）
                    from openpyxl.utils import get_column_letter
                    from openpyxl.worksheet.page import PageMargins

                    print_area = "$A$1:$F$39"
                    ws.print_area = print_area

                    # 印刷位置（マージン）を設定（適度な余白を確保）
                    # E列が別ページになるのを防ぐため、左右マージンを少し小さく設定
                    # 中央寄せのために左右のマージンを同じ値に設定
                    ws.page_margins = PageMargins(
                        left=0.6,  # 左マージン（インチ）-
                        right=0.4,  # 右マージン（インチ）-
                        top=0.5,  # 上マージン（インチ）
                        bottom=0.5,  # 下マージン（インチ）
                        header=0.3,  # ヘッダーマージン（インチ）
                        footer=0.3,  # フッターマージン（インチ）
                    )

                    # ページ設定を調整（縦向きA4、横幅を1ページに収める）
                    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # 縦向き
                    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # A4サイズ
                    ws.page_setup.fitToWidth = 1  # 横幅を1ページに収める
                    ws.page_setup.fitToHeight = 0  # 高さは無制限（自動調整）
                    ws.page_setup.horizontalCentered = True  # 水平方向の中央寄せ
                    ws.page_setup.verticalCentered = False  # 垂直方向は中央寄せしない

                    # セルの幅と高さを調整して1ページに収める（小さくしすぎない）
                    # A4縦向きの印刷可能幅: 約8.27インチ - 左右マージン(0.8インチ) = 7.47インチ
                    # Excelの列幅は文字数単位（標準フォント11ptで約7ピクセル/文字）
                    # 1インチ = 96ピクセル、印刷可能幅 ≈ 7.47 * 96 / 7 ≈ 102文字分
                    # ただし、小さくしすぎないように90%程度に調整

                    # 現在の列幅の合計を計算（A列からF列まで）
                    total_width = 0
                    for col_idx in range(1, 7):  # A列(1)からF列(6)まで
                        col_letter = get_column_letter(col_idx)
                        col_dim = ws.column_dimensions.get(col_letter)
                        if col_dim and col_dim.width:
                            total_width += col_dim.width
                        else:
                            total_width += 8.43  # デフォルト列幅

                    # A4縦向きの印刷可能幅（文字数）: 約102文字（マージン0.4インチを考慮）
                    # 6列分（A列からF列）を考慮して、より適切な幅に調整
                    # 小さくしすぎないように、最大90%まで縮小
                    max_printable_width = 100
                    if total_width > max_printable_width:
                        # スケールファクターを計算（最小0.5、最大0.9）
                        scale_factor = min(0.9, max(0.4, max_printable_width / total_width))
                        # 列幅を調整（A列からF列まで）
                        for col_idx in range(1, 7):  # A列(1)からF列(6)まで
                            col_letter = get_column_letter(col_idx)
                            col_dim = ws.column_dimensions[col_letter]
                            if col_dim.width:
                                col_dim.width = col_dim.width * scale_factor
                            else:
                                col_dim.width = 8.43 * scale_factor

                    # 行の高さも調整（必要に応じて）
                    # A4縦向きの印刷可能高さ: 約11.69インチ - 上下マージン(1.0インチ) = 10.69インチ
                    # 1ポイント = 1/72インチ、印刷可能高さ ≈ 10.69 * 72 ≈ 770ポイント
                    # 小さくしすぎないように、最大90%まで縮小

                    # 現在の行の高さの合計を計算（1行目から39行目まで）
                    total_height = 0
                    for row_idx in range(1, 40):  # 1行目から39行目まで
                        row_dim = ws.row_dimensions.get(row_idx)
                        if row_dim and row_dim.height:
                            total_height += row_dim.height
                        else:
                            total_height += 15  # デフォルト行の高さ（ポイント）

                    # A4縦向きの印刷可能高さ（ポイント）: 約770ポイント
                    max_printable_height = 770
                    if total_height > max_printable_height:
                        # スケールファクターを計算（最小0.5、最大0.9）
                        height_scale_factor = min(0.9, max(0.5, max_printable_height / total_height))
                        # 行の高さを調整（1行目から39行目まで）
                        for row_idx in range(1, 40):  # 1行目から39行目まで
                            row_dim = ws.row_dimensions[row_idx]
                            if row_dim.height:
                                row_dim.height = row_dim.height * height_scale_factor
                            else:
                                row_dim.height = 15 * height_scale_factor
            # 一時ファイルとして保存
            temp_excel_path = excel_path.replace(".xlsx", "_temp_pdf.xlsx")
            wb.save(temp_excel_path)
            wb.close()
            excel_file_to_convert = temp_excel_path
        except Exception as e:
            print(f"印刷範囲の設定中にエラーが発生しました: {str(e)}")
            # エラーが発生した場合は元のファイルを使用
            excel_file_to_convert = excel_path

        # LibreOfficeがインストールされているか確認
        libreoffice_path = shutil.which("libreoffice")
        if not libreoffice_path:
            # macOSの場合、一般的なパスを確認
            possible_paths = [
                "/Applications/LibreOffice.app/Contents/MacOS/soffice",
                "/usr/local/bin/libreoffice",
                "/opt/homebrew/bin/libreoffice",
            ]
            libreoffice_path = None
            for path in possible_paths:
                if os.path.exists(path):
                    libreoffice_path = path
                    break

            if not libreoffice_path:
                return (
                    None,
                    "LibreOfficeがインストールされていません。"
                    "macOSの場合: brew install --cask libreoffice でインストールしてください。",
                )

        # PDFファイルのパスを生成
        pdf_path = os.path.splitext(excel_path)[0] + ".pdf"
        output_dir = os.path.dirname(excel_path)

        # LibreOfficeコマンドでPDFに変換
        # --headless: GUIなしで実行
        # --convert-to pdf: PDF形式に変換（各シートが1ページとして出力される）
        # --outdir: 出力ディレクトリを指定
        # calc_pdf_Export: Calc（Excel）用のPDFエクスポートフィルターを使用
        command = [
            libreoffice_path,
            "--headless",
            "--convert-to",
            "pdf:calc_pdf_Export",
            "--outdir",
            output_dir,
            excel_file_to_convert,
        ]

        try:
            # コマンドを実行
            result = subprocess.run(
                command,
                capture_output=True,
                text=True,
                timeout=60,  # 60秒のタイムアウト
            )

            # 一時ファイルを削除
            if temp_excel_path and os.path.exists(temp_excel_path):
                try:
                    os.remove(temp_excel_path)
                except Exception:
                    pass

            if result.returncode == 0:
                # PDFファイルが生成されたか確認
                # 一時ファイルを使用した場合、PDFファイル名も調整が必要
                if temp_excel_path:
                    temp_pdf_path = os.path.splitext(temp_excel_path)[0] + ".pdf"
                    if os.path.exists(temp_pdf_path):
                        # 元のファイル名にリネーム
                        if os.path.exists(pdf_path):
                            os.remove(pdf_path)
                        os.rename(temp_pdf_path, pdf_path)

                if os.path.exists(pdf_path):
                    print(f"PDFファイルを生成しました: {pdf_path}")
                    return pdf_path, None
                else:
                    # LibreOfficeは元のファイル名を維持するため、パスを確認
                    # ファイル名にスペースが含まれている場合の処理
                    base_name = os.path.splitext(os.path.basename(excel_path))[0]
                    possible_pdf_path = os.path.join(output_dir, f"{base_name}.pdf")
                    if os.path.exists(possible_pdf_path):
                        return possible_pdf_path, None
                    return None, "PDFファイルが生成されませんでした"
            else:
                error_msg = f"LibreOffice変換エラー: {result.stderr}"
                print(error_msg)
                return None, error_msg

        except subprocess.TimeoutExpired:
            return None, "PDF変換がタイムアウトしました（60秒以上かかりました）"
        except Exception as e:
            error_msg = f"PDF変換中にエラーが発生しました: {str(e)}"
            print(error_msg)
            return None, error_msg

    except Exception as e:
        error_msg = f"PDF変換中にエラーが発生しました: {str(e)}"
        print(error_msg)
        import traceback

        traceback.print_exc()
        return None, error_msg


def _extract_year_month_from_sales(sales_filename: str) -> tuple[int | None, int | None]:
    """
    売上データファイル（Excel）から年と月を自動で抽出する内部関数

    Args:
        sales_filename (str): 売上データのファイル名

    Returns:
        tuple[int | None, int | None]: (年, 月) のタプル。抽出に失敗した場合は (None, None)
    """
    try:
        # このファイルがあるディレクトリの絶対パスを取得
        base_dir = os.path.dirname(os.path.abspath(__file__))
        # 売上データのフルパスを作成
        sales_file = os.path.join(base_dir, "uploads", "sales", sales_filename)

        # ファイルが存在しない場合は、Noneを返す
        if not os.path.exists(sales_file):
            return None, None

        # pandasでExcelファイルを読み込む
        sales_df = pd.read_excel(sales_file)

        # "売上日"という列が存在しない場合は、Noneを返す
        if "売上日" not in sales_df.columns:
            return None, None

        # "売上日"の列を日付型に変換（時刻は切り捨てる）
        sales_df["売上日"] = pd.to_datetime(sales_df["売上日"]).dt.date

        # "売上日"列が空の場合は、Noneを返す
        if sales_df["売上日"].empty:
            return None, None

        # 最も古い日付と最も新しい日付を取得
        min_date = sales_df["売上日"].min()
        max_date = sales_df["売上日"].max()

        # 全てのデータが同じ月のものであれば、その年月を返す
        if min_date.year == max_date.year and min_date.month == max_date.month:
            return min_date.year, min_date.month

        # 複数月にまたがるデータの場合は、最も新しい月を対象とする
        return max_date.year, max_date.month

    except Exception:
        # 何らかのエラーが発生した場合（例：ファイル形式が不正など）はNoneを返す
        return None, None


# モデルの定義（循環インポートを避けるため、ここで定義）
class SettlementHistory(db.Model):
    """
    精算書発行履歴テーブル

    精算書の一括生成・ダウンロード履歴を保存する。

    Attributes:
        id: 主キー（自動採番）
        year: 精算対象年
        month: 精算対象月
        file_name: 生成されたファイル名
        file_path: ファイルの保存パス
        file_format: ファイル形式（excel/pdf）
        created_at: 発行日時（自動設定）
    """

    __tablename__ = "settlement_history"

    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    year = db.Column(db.Integer, nullable=False, comment="精算対象年")
    month = db.Column(db.Integer, nullable=False, comment="精算対象月")
    file_name = db.Column(db.String(255), nullable=False, comment="ファイル名")
    file_path = db.Column(db.String(500), nullable=False, comment="ファイル保存パス")
    file_format = db.Column(db.String(10), nullable=False, default="excel", comment="ファイル形式")
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, comment="発行日時")

    # インデックスの定義
    __table_args__ = (
        db.Index("idx_settlement_history_year_month", "year", "month"),
        db.Index("idx_settlement_history_created_at", "created_at"),
    )

    def __repr__(self) -> str:
        """オブジェクトの文字列表現を返す"""
        return f"<SettlementHistory {self.id}: {self.year}年{self.month}月 - {self.file_name}>"

    def to_dict(self) -> dict:
        """辞書形式に変換"""
        return {
            "id": self.id,
            "year": self.year,
            "month": self.month,
            "file_name": self.file_name,
            "file_path": self.file_path,
            "file_format": self.file_format,
            "created_at": self.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        }


# --- ルート（URL）定義 ---


@app.route("/", methods=["GET"])
def index():
    """
    トップページ（/）にアクセスされたときに呼ばれる関数
    ファイルアップロード用のHTML画面を表示します。
    """
    # "templates"フォルダの中の"upload.html"を読み込んでユーザーに表示
    return render_template("upload.html")


@app.route("/history", methods=["GET"])
def history_page():
    """
    発行履歴一覧ページを表示
    """
    # 発行履歴を新しい順に取得
    histories = SettlementHistory.query.order_by(SettlementHistory.created_at.desc()).all()

    return render_template("history.html", histories=histories)


@app.route("/download/<int:history_id>")
def download_page(history_id):
    """
    精算書ダウンロードページを表示
    """
    history = SettlementHistory.query.get_or_404(history_id)

    # ファイルの存在確認
    if not os.path.exists(history.file_path):
        flash(f"ファイルが見つかりません: {history.file_name}", "error")
        return redirect("/")

    return render_template("download.html", history=history)


@app.route("/download/<int:history_id>/file")
def download_file(history_id):
    """
    精算書ファイルをダウンロード（形式を指定）
    """
    history = SettlementHistory.query.get_or_404(history_id)
    file_format = request.args.get("format", "excel")  # excel または pdf

    # ファイルの存在確認
    if not os.path.exists(history.file_path):
        flash(f"ファイルが見つかりません: {history.file_name}", "error")
        return redirect("/")

    if file_format == "excel":
        # Excelファイルをそのままダウンロード
        return send_from_directory(
            os.path.dirname(history.file_path), os.path.basename(history.file_path), as_attachment=True
        )
    elif file_format == "pdf":
        # PDFに変換してダウンロード
        try:
            pdf_path, error_message = _convert_excel_to_pdf(history.file_path)
            if pdf_path and os.path.exists(pdf_path):
                return send_from_directory(
                    os.path.dirname(pdf_path), os.path.basename(pdf_path), as_attachment=True
                )
            else:
                # エラーメッセージがあればそれを表示、なければデフォルトメッセージ
                if error_message:
                    flash(f"PDF変換に失敗しました: {error_message}", "error")
                else:
                    flash(
                        "PDF変換に失敗しました。LibreOfficeが正しくインストールされているか確認してください。",
                        "error",
                    )
                return redirect(f"/download/{history_id}")
        except Exception as e:
            flash(f"PDF変換中にエラーが発生しました: {str(e)}", "error")
            return redirect(f"/download/{history_id}")
    else:
        flash("無効なファイル形式が指定されました", "error")
        return redirect(f"/download/{history_id}")


@app.route("/history/<int:history_id>/download")
def download_history(history_id):
    """
    履歴から過去の精算書をダウンロード（形式を指定）
    """
    history = SettlementHistory.query.get_or_404(history_id)
    file_format = request.args.get("format", "excel")  # excel または pdf

    # ファイルの存在確認
    if not os.path.exists(history.file_path):
        flash(f"ファイルが見つかりません: {history.file_name}", "error")
        return redirect("/history")

    if file_format == "excel":
        # Excelファイルをそのままダウンロード
        return send_from_directory(
            os.path.dirname(history.file_path), os.path.basename(history.file_path), as_attachment=True
        )
    elif file_format == "pdf":
        # PDFに変換してダウンロード
        try:
            pdf_path, error_message = _convert_excel_to_pdf(history.file_path)
            if pdf_path and os.path.exists(pdf_path):
                return send_from_directory(
                    os.path.dirname(pdf_path), os.path.basename(pdf_path), as_attachment=True
                )
            else:
                # エラーメッセージがあればそれを表示、なければデフォルトメッセージ
                if error_message:
                    flash(f"PDF変換に失敗しました: {error_message}", "error")
                else:
                    flash(
                        "PDF変換に失敗しました。LibreOfficeが正しくインストールされているか確認してください。",
                        "error",
                    )
                return redirect("/history")
        except Exception as e:
            flash(f"PDF変換中にエラーが発生しました: {str(e)}", "error")
            return redirect("/history")
    else:
        flash("無効なファイル形式が指定されました", "error")
        return redirect("/history")


@app.route("/history/<int:history_id>/delete", methods=["POST"])
def delete_history(history_id):
    """
    発行履歴を削除する
    """
    history = SettlementHistory.query.get_or_404(history_id)

    try:
        # ファイルが存在する場合は削除
        if os.path.exists(history.file_path):
            try:
                os.remove(history.file_path)
                # PDFファイルも存在する場合は削除
                pdf_path = os.path.splitext(history.file_path)[0] + ".pdf"
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
            except Exception as e:
                print(f"ファイル削除エラー: {str(e)}")
                # ファイル削除に失敗してもデータベースレコードは削除する

        # データベースから削除
        file_name = history.file_name
        db.session.delete(history)
        db.session.commit()

        flash(f"発行履歴「{file_name}」を削除しました。", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"削除に失敗しました: {str(e)}", "error")

    return redirect("/history")


@app.route("/upload", methods=["POST"])
def upload_files():
    """
    ファイルアップロード処理のメインとなる関数
    顧客データと売上データの両方がアップロードされたら、自動で精算書を生成します。
    """
    customer_filename = None  # アップロードされた顧客データファイル名を保存する変数
    sales_filename = None  # アップロードされた売上データファイル名を保存する変数

    # --- 顧客データの処理 ---
    # "customer_file"という名前のファイルがフォームから送信されたかチェック
    if "customer_file" in request.files:
        customer_file = request.files["customer_file"]
        # ファイル名が空でないことを確認
        if customer_file.filename != "":
            # ファイルが存在し、かつ許可された拡張子かチェック
            if customer_file and allowed_file(customer_file.filename):
                try:
                    # 保存先フォルダのパスを作成
                    customers_folder = os.path.join(app.config["UPLOAD_FOLDER"], "customers")
                    # フォルダがなければ作成
                    if not os.path.exists(customers_folder):
                        os.makedirs(customers_folder)
                    # ファイル名を安全な形式に変換
                    filename = secure_filename(customer_file.filename)
                    # ファイルを保存
                    customer_file.save(os.path.join(customers_folder, filename))
                    customer_filename = filename  # ファイル名を後で使うために保存
                    # 成功メッセージをflashで表示
                    flash(f"顧客データ: {filename} がアップロードされました", "success")
                except Exception as e:
                    # エラーメッセージをflashで表示
                    flash(
                        f"顧客データ: {customer_file.filename} のアップロードに失敗しました: {str(e)}",
                        "error",
                    )
            else:
                # 許可されていないファイル形式の場合のエラーメッセージ
                flash(
                    f"顧客データ: {customer_file.filename} は許可されていないファイル形式です（Excelファイルのみ許可）",
                    "error",
                )

    # --- 委託販売売上データの処理 (顧客データと同様の流れ) ---
    if "sales_file" in request.files:
        sales_file = request.files["sales_file"]
        if sales_file.filename != "":
            if sales_file and allowed_file(sales_file.filename):
                try:
                    sales_folder = os.path.join(app.config["UPLOAD_FOLDER"], "sales")
                    if not os.path.exists(sales_folder):
                        os.makedirs(sales_folder)
                    filename = secure_filename(sales_file.filename)
                    sales_file.save(os.path.join(sales_folder, filename))
                    sales_filename = filename
                    flash(f"委託販売売上データ: {filename} がアップロードされました", "success")
                except Exception as e:
                    flash(
                        f"委託販売売上データ: {sales_file.filename} のアップロードに失敗しました: {str(e)}",
                        "error",
                    )
            else:
                flash(
                    f"委託販売売上データ: {sales_file.filename} は許可されていないファイル形式です（Excelファイルのみ許可）",
                    "error",
                )

    # --- ファイル選択のバリデーション ---
    customer_selected = "customer_file" in request.files and request.files["customer_file"].filename != ""
    sales_selected = "sales_file" in request.files and request.files["sales_file"].filename != ""
    # どちらのファイルも選択されずにフォームが送信された場合
    if not customer_selected and not sales_selected:
        flash("少なくとも1つのファイルを選択してください", "error")
        return redirect("/")  # トップページに戻る

    # --- 精算書の自動生成処理 ---
    # 顧客データと売上データの両方が正常にアップロードされた場合に実行
    if customer_filename and sales_filename:
        try:
            # 売上データファイルから自動で年月を抽出
            year, month = _extract_year_month_from_sales(sales_filename)

            # 年月が抽出できなかった場合
            if not year or not month:
                flash(
                    "売上データから年と月を抽出できませんでした。売上データに有効な日付が含まれている必要があります。",
                    "error",
                )
                return redirect("/")

            # --- 精算書生成に必要なファイルパスを準備 ---
            base_dir = os.path.dirname(os.path.abspath(__file__))
            customer_file = os.path.join(base_dir, "uploads", "customers", customer_filename)
            sales_file = os.path.join(base_dir, "uploads", "sales", sales_filename)
            template_file = os.path.join(base_dir, "settlement_template.xlsx")
            output_dir = os.path.join(base_dir, "outputs")
            os.makedirs(output_dir, exist_ok=True)  # 出力先フォルダがなければ作成

            # 各ファイルが実際に存在するか確認
            if not os.path.exists(customer_file):
                flash(f"顧客データファイルが見つかりません: {customer_filename}", "error")
                return redirect("/")
            if not os.path.exists(sales_file):
                flash(f"売上データファイルが見つかりません: {sales_filename}", "error")
                return redirect("/")
            if not os.path.exists(template_file):
                flash("テンプレートファイルが見つかりません: settlement_template.xlsx", "error")
                return redirect("/")

            # --- 精算書生成の実行 ---
            # 別のファイル(services/settlement_generator.py)に定義された関数を呼び出す
            output_filename = create_settlements_for_month(
                year=year,
                month=month,
                customer_file=customer_file,
                sales_file=sales_file,
                template_file=template_file,
                output_dir=output_dir,
            )

            # 生成結果に応じてメッセージを表示
            if output_filename:
                output_basename = os.path.basename(output_filename)  # パスからファイル名だけを取得

                # 発行履歴をデータベースに保存
                history = SettlementHistory(
                    year=year,
                    month=month,
                    file_name=output_basename,
                    file_path=output_filename,
                    file_format="excel",
                )
                db.session.add(history)
                db.session.commit()

                # ダウンロードページにリダイレクト
                return redirect(f"/download/{history.id}")
            else:
                flash("精算書の生成に失敗しました。対象期間のデータが存在しない可能性があります。", "error")

        except ValueError as e:
            # 想定される入力値エラー（例：数値変換失敗など）
            flash(f"入力値が不正です: {str(e)}", "error")
        except Exception as e:
            # その他の予期せぬエラー
            flash(f"精算書の自動生成に失敗しました: {str(e)}", "error")

    # 処理が終わったらトップページにリダイレクト
    return redirect("/")


@app.route("/upload/customers", methods=["POST"])
def upload_customers():
    """
    【現在不使用の可能性】顧客データファイル単体をアップロードするための関数
    （メインの/uploadルートで一括処理しているため、通常は使われない）
    """
    if "file" not in request.files:
        flash("ファイルが選択されていません", "error")
        return redirect("/")

    file = request.files["file"]

    if file.filename == "":
        flash("ファイルが選択されていません", "error")
        return redirect("/")

    customers_folder = os.path.join(app.config["UPLOAD_FOLDER"], "customers")
    if not os.path.exists(customers_folder):
        os.makedirs(customers_folder)

    if file and allowed_file(file.filename):
        try:
            filename = secure_filename(file.filename)
            file.save(os.path.join(customers_folder, filename))
            flash(f"顧客データ: {filename} がアップロードされました", "success")
        except Exception:
            flash(f"顧客データ: {file.filename} のアップロードに失敗しました", "error")
    else:
        flash(
            f"顧客データ: {file.filename} は許可されていないファイル形式です（Excelファイルのみ許可）",
            "error",
        )

    return redirect("/")


@app.route("/upload/sales", methods=["POST"])
def upload_sales():
    """
    【現在不使用の可能性】売上データファイル単体をアップロードするための関数
    （メインの/uploadルートで一括処理しているため、通常は使われない）
    """
    if "file" not in request.files:
        flash("ファイルが選択されていません", "error")
        return redirect("/")

    file = request.files["file"]

    if file.filename == "":
        flash("ファイルが選択されていません", "error")
        return redirect("/")

    sales_folder = os.path.join(app.config["UPLOAD_FOLDER"], "sales")
    if not os.path.exists(sales_folder):
        os.makedirs(sales_folder)

    if file and allowed_file(file.filename):
        try:
            filename = secure_filename(file.filename)
            file.save(os.path.join(sales_folder, filename))
            flash(f"委託販売売上データ: {filename} がアップロードされました", "success")
        except Exception:
            flash(f"委託販売売上データ: {file.filename} のアップロードに失敗しました", "error")
    else:
        flash(
            f"委託販売売上データ: {file.filename} は許可されていないファイル形式です（Excelファイルのみ許可）",
            "error",
        )

    return redirect("/")


@app.route("/uploads/<filename>")
def uploaded_file(filename):
    """
    アップロードされたファイルをブラウザで表示・ダウンロードさせるための関数
    （例：/uploads/my_data.xlsx にアクセスするとファイルが返される）
    ※セキュリティ上、直接ファイルシステムにアクセスさせないためにこの関数が存在します
    """
    # send_from_directoryを使うことで、指定されたディレクトリ外のファイルへのアクセスを防ぐ
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


@app.route("/generate", methods=["POST"])
def generate_settlements():
    """
    【現在不使用の可能性】手動で精算書を生成するための関数
    （/uploadルートでの自動生成がメインのため、通常は使われない）
    """
    try:
        # フォームから年、月、ファイル名を取得
        year = int(request.form.get("year"))
        month = int(request.form.get("month"))
        customer_filename = request.form.get("customer_file")
        sales_filename = request.form.get("sales_file")

        # バリデーション（入力値のチェック）
        if not customer_filename or not sales_filename:
            flash("顧客データファイルと売上データファイルの両方を指定してください", "error")
            return redirect("/")
        if not (1 <= month <= 12):
            flash("月は1から12の間で指定してください", "error")
            return redirect("/")

        # ファイルパスの構築
        base_dir = os.path.dirname(os.path.abspath(__file__))
        customer_file = os.path.join(base_dir, "uploads", "customers", customer_filename)
        sales_file = os.path.join(base_dir, "uploads", "sales", sales_filename)
        template_file = os.path.join(
            base_dir, "委託販売精算書.xlsx"
        )  # 注意：テンプレート名がハードコードされている

        # ファイルの存在確認
        if not os.path.exists(customer_file):
            flash(f"顧客データファイルが見つかりません: {customer_filename}", "error")
            return redirect("/")
        if not os.path.exists(sales_file):
            flash(f"売上データファイルが見つかりません: {sales_filename}", "error")
            return redirect("/")
        if not os.path.exists(template_file):
            flash("テンプレートファイルが見つかりません: 委託販売精算書.xlsx", "error")
            return redirect("/")

        # 出力ディレクトリの準備
        output_dir = os.path.join(base_dir, "outputs")
        os.makedirs(output_dir, exist_ok=True)

        # 精算書生成の実行
        output_filename = create_settlements_for_month(
            year=year,
            month=month,
            customer_file=customer_file,
            sales_file=sales_file,
            template_file=template_file,
            output_dir=output_dir,
        )

        # 結果の表示
        if output_filename:
            output_basename = os.path.basename(output_filename)
            flash(f"精算書の生成が完了しました: {output_basename}", "success")
        else:
            flash("精算書の生成に失敗しました。対象期間のデータが存在しない可能性があります。", "error")

    except ValueError as e:
        flash(f"入力値が不正です: {str(e)}", "error")
    except Exception as e:
        flash(f"精算書の生成に失敗しました: {str(e)}", "error")

    return redirect("/")


@app.route("/outputs/<filename>")
def download_output(filename):
    """
    生成された精算書ファイルをダウンロードさせるための関数
    （例：/outputs/2025年10月度_精算書.xlsx にアクセスするとダウンロードが始まる）
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(base_dir, "outputs")
    # send_from_directoryで安全にファイルを送信
    return send_from_directory(output_dir, filename)


# --- アプリケーションの実行 ---

# このファイルが `python app.py` のように直接実行された場合にのみ、以下のコードが実行される
if __name__ == "__main__":
    """
    開発用Webサーバーを起動します。
    """
    # データベーステーブルの作成
    with app.app_context():
        db.create_all()

    # 起動前に、アップロード用のフォルダが存在するか確認し、なければ作成する
    # これにより、初回起動時にアップロードでエラーが出るのを防ぎます
    if not os.path.exists(app.config["UPLOAD_FOLDER"]):
        os.makedirs(app.config["UPLOAD_FOLDER"])

    # Flaskに組み込まれている開発用サーバーを起動
    # debug=True: デバッグモードを有効にする。
    #   - コードを変更するとサーバーが自動で再起動する
    #   - エラーが発生した際に、ブラウザ上で詳細なエラー情報を確認できる
    debug = os.environ.get("FLASK_DEBUG", "False").lower() == "true"

    # 標準出力のバッファリングを無効化（Windows環境で確実に表示されるように）
    if sys.platform == "win32":
        import io

        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", line_buffering=True)
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", line_buffering=True)

    app.run(debug=debug, host="127.0.0.1", port=5000)
